import streamlit as st
import json
import os
import pandas as pd
import glob, os.path
from datetime import date, datetime, timedelta
from dateutil.relativedelta import relativedelta
import requests

import matplotlib.pyplot as plt
from matplotlib.dates import YearLocator, MonthLocator, DayLocator, HourLocator, MinuteLocator, DateFormatter
from matplotlib.patches import Rectangle
import matplotlib.patches as patches
from matplotlib.patches import Rectangle
import matplotlib.patches as patches
import numpy as np
#import seaborn as sns

import altair as alt
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import plotly.io
from bokeh.plotting import figure

#doc
import openpyxl
from openpyxl.styles import Font, Color, Alignment, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.drawing.image import Image
from io import BytesIO





class Plot():
    
    def __init__(self, estacion,sensor, fecha_ini,fecha_fin,list_est):

        #self.folder = download_folder
        self.estacion = estacion
        self.sensor = sensor
        self.fecha_ini = fecha_ini
        self.fecha_fin = fecha_fin
        self.list_est = list_est
        
        df_est = pd.read_csv(list_est[0])
        #st.dataframe(df_est)
        
        #filtro de tiempo
        
        #Cambio de formato de fecha a datetime
        f_date = []
        for e in df_est["fecha"]:
            year, month, day = e.split("-")
            f_date.append(datetime(int(year),int(month),int(day)))
        
        #Cambio de formato de fecha times_gaps a datetime
        df_est["fecha"] = f_date
        
        f_date_gap = []
        f_date_gap_graf = []
        f_date_gap_graf_din = []
        for te, dis, fe in zip(df_est["times_gaps"], df_est["disponibilidad"], df_est["fecha"]):
            
            
            if float(dis) == 0.0:
                f_date_gap.append([[datetime(fe.year,fe.month,fe.day,0,0,0,0),3600*24]])
                
                f_date_gap_graf.append([[datetime(fe.year,fe.month,fe.day,0,0,0,0),3600*24]])
                f_date_gap_graf_din.append([[datetime(fe.year,fe.month,fe.day,0,0,0,0),datetime(fe.year,fe.month,fe.day,23,59,59)]])


            if len(te) != 2 and float(dis) != 0.0:
                t_gaps = te.split("_")

                gaps_dia = []    
                gaps_dia_din = []                
                for t in t_gaps:

                    if len(t) != 0:
                        y_g, mo_g, d_g, h_g, mi_g, s_g, ms_g, del_g = t.split("-")
                        gaps_dia.append([datetime(int(y_g),int(mo_g),int(d_g),int(h_g),int(mi_g),int(s_g),int(ms_g)),float(del_g)])
                        gaps_dia_din.append([datetime(int(y_g),int(mo_g),int(d_g),int(h_g),int(mi_g),int(s_g),int(ms_g)),datetime(int(y_g),int(mo_g),int(d_g),int(h_g),int(mi_g),int(s_g),int(ms_g))+timedelta(seconds=float(del_g))])


                f_date_gap.append(gaps_dia)
                f_date_gap_graf.append(gaps_dia)
                f_date_gap_graf_din.append(gaps_dia_din)
            
            elif len(te) == 2 and dis != 0.0:
                #f_date_gap.append([[datetime(fe.year,fe.month,fe.day,0,0,0,0),0]])
                f_date_gap.append([[[],[]]])
                f_date_gap_graf_din.append([[[],[]]])
                
        
        #st.write(f_date_gap_graf_din)
        #remontar fechas con datetime
    
        df_est["times_gaps"]= f_date_gap
        

        #cree una nueva columna para los datos de tiempos de gaps para el mapa dinamico
        df_est["times_gaps_graf_din"]= f_date_gap_graf_din
        

            
        
        #filtro por fechas seleccionadas
        filtro_fecha = (df_est["fecha"] >= fecha_ini) & (df_est["fecha"] <= fecha_fin) 
        fil_time = df_est[filtro_fecha]
        
        
        del_time_m=(fecha_fin-fecha_ini).total_seconds()/60/60/24/30  #en meses
        dias_ad=int(round((4*del_time_m)-3 ,0 ) ) #al hacerce mas larga la grafica, debe adicionar mas dias para construir columna de calor (para que se vea)
        
        if dias_ad <= 0:
            dias_ad=1

        self.fil_time = fil_time
        self.dias_ad = dias_ad
            
    
    
    def plot_disp_fij(self, can, folder,two_sensores=False):
        
        #folder = self.folder
        estacion = self.estacion
        sensor = self.sensor
        fecha_ini = self.fecha_ini
        fecha_fin = self.fecha_fin
        fil_time = self.fil_time
        dias_ad = self.dias_ad  
        list_est= self.list_est 
        
        if len(fil_time) != 0:
            
            cod_sen=""
            if two_sensores==True:
                cod_sen=list_est[0][-10:-8]
            
            
            #-----GRAFICA DISPONIBILIDAD
            
            #fil_time[" disponibilidad"] = fil_time[" disponibilidad"].astype(float)
            fil_time["disponibilidad"] = list(map(lambda x: float(x), fil_time["disponibilidad"]))
            
            lista_disponibilidad_sin_nueves=[] #es posible que no haya podido leer un mseed por tanto escribio un -9
            for di in fil_time["disponibilidad"]:
                if di != -9:
                    lista_disponibilidad_sin_nueves.append(di)

            fil_time2 = fil_time[(fil_time["disponibilidad"] != -9)]
            min_dis = fil_time2["disponibilidad"].min()             #minima desponibilidad  
            max_dis = fil_time2["disponibilidad"].max()             #maxima disponibilidad
            prom_dis = fil_time2["disponibilidad"].mean()           #promedio disponibilidad
            
            #min_dis=round( min(lista_disponibilidad_sin_nueves),1 )
            #max_dis=round( max(lista_disponibilidad_sin_nueves),1 )
            #prom_dis=round( np.mean( np.array(lista_disponibilidad_sin_nueves) ),1 )
            
            #col1,col2= st.columns([5,1.1])
            #with col1:
            #Con Matplotlib
            fig = plt.figure(figsize=(15 , 8))
            ax = fig.add_subplot(311)
            plt.plot(fil_time["fecha"],fil_time["disponibilidad"],linewidth=1,color='#0000ff',label="Disponibilidad",zorder=1000)
            ax.set_ylabel('Disponibilidad \n (%)', color='k', fontsize=14)
            ax.xaxis.grid(True, which='major')
            ax.yaxis.grid(True, which='major')
            ax.xaxis.set_tick_params(labelsize=9)
            #ax.set_xlim(list(fil_time["fecha"])[0]-timedelta(days=1), list(fil_time["fecha"])[-1]+timedelta(days=dias_ad)) #AQUÍ ES PARA AGREGAR EL ESPACIO PARA AGREGAR EL RECTANGULO DE PROYECCIÓN DE GAPS
            ax.set_xlim(list(fil_time["fecha"])[0]-timedelta(days=1), list(fil_time["fecha"])[-1])

            #con este foorloop se selecciona en una lista solo los tiempos de gaps existentes
            l_fil_tgap = []
            for fil_time_gap in fil_time["times_gaps"]:
                if len(str(fil_time_gap[0][0])) > 2 :
                    l_fil_tgap.append(fil_time_gap)
            
                    
            if len(l_fil_tgap) > 0:
            
                ax1 = ax.twinx()
                f_i = datetime(fecha_ini.year,fecha_ini.month,fecha_ini.day)
                f_f = datetime(fecha_fin.year,fecha_fin.month,fecha_fin.day,0,0,0,0)
                ax1.vlines(f_i,0,0, lw=1, color="k",label="Gaps (tiempo)" )  
                to_r=datetime(1,1,1,0,0,1)
                #rectangulo3 = patches.Rectangle( ( list(fil_time["fecha"])[-1], to_r ),timedelta(dias_ad),timedelta(hours=24),fill=True,facecolor='b',edgecolor='b',linewidth=0.2,alpha=0.6)
                #ax1.add_patch(rectangulo3)   #CON ESTAS DOS LINEAS SE AGREGA EL RECTANGULO AZUL DEL FINAL DEL GRAFICO PARA PROLLECTAR LOS GAPS
                
                #para graficar los gaps existentes
                for tg in l_fil_tgap:
                    for g in tg:
                        tdia=datetime(g[0].year,g[0].month,g[0].day,12,0,0)
                        to=datetime(1,1,1,g[0].hour,g[0].minute,g[0].second)
                        ax1.text(tdia,to,"o",fontsize=7,alpha=0.8,zorder=100,horizontalalignment='center',verticalalignment='baseline')
                        ax1.text(tdia,to+timedelta(seconds=g[1]),"x",fontsize=10,alpha=0.8,zorder=100,horizontalalignment='center',verticalalignment='baseline')
                        rectangulo = patches.Rectangle( ( tdia-timedelta(seconds=int(g[1]/2)), to),timedelta(seconds=g[1]),timedelta(seconds=g[1]),fill=True,facecolor='k',edgecolor='k',linewidth=0.6,alpha=0.6, ) 
                        ax1.add_patch(rectangulo)
        
                        if g[1] < 3600*24:#si el gap fue de todo el dia no lo grafica en columna calor
                            rectangulo2 = patches.Rectangle( ( list(fil_time["fecha"])[-1], to),timedelta(dias_ad),timedelta(seconds=g[1]),fill=True,facecolor='k',edgecolor='k',linewidth=0.2,alpha=0.08, ) 
                            ax1.add_patch(rectangulo2)
                
                
                ax1.set_ylim( datetime(1,1,1,0,0,0), datetime(1,1,1,23,59,59))
                ax1.yaxis.set_minor_locator( HourLocator(interval = int(2)))
                ax1.yaxis.set_minor_formatter( DateFormatter('%H:%m') )
                ax1.yaxis.set_major_locator( HourLocator(interval = int(8)))
                ax1.yaxis.set_major_formatter( DateFormatter('%H') )  
                
                labels = ax1.yaxis.get_minorticklabels()
                plt.setp(labels, rotation=0, fontsize=10)
                labels = ax1.get_yticklabels() 
                plt.setp(labels, rotation=0, fontsize=9) 
                
                ax1.set_ylabel('Gaps\nHora (UTC)', color='k', fontsize=14)
                ax1.legend(loc='upper right')
            
            ax.legend(loc='upper left')        
            plt.title('Disponibilidad ' +estacion+" "+sensor+can+" | "+fecha_ini.strftime("%Y-%m-%d")+" - "+fecha_fin.strftime("%Y-%m-%d"))

            plt.savefig(folder+f"disp_{estacion}_{sensor}{can}{cod_sen}.png", bbox_inches='tight')

            #st.pyplot(fig=fig, )
                
            #with col2:
                
                #st.markdown(f"**:blue[Min. disp.:]** {min_dis}%**:blue[| Max. disp.:]** {max_dis}%")
                #st.write(f"Max. disp: {max_dis}")
                #st.markdown(f"**:blue[Promedio disp.:]** {round(prom_dis,2)}%")
                #st.write("_________________________")
        else:
            print(f"No hay datos para la estación {estacion} entre las fechas del {fecha_ini} {fecha_fin}")        
    
    def plot_off_fij(self, can, folder,two_sensores=False):
        
        #folder = self.folder
        estacion = self.estacion
        sensor = self.sensor
        fecha_ini = self.fecha_ini
        fecha_fin = self.fecha_fin
        fil_time = self.fil_time
        dias_ad = self.dias_ad  
        list_est= self.list_est
         
        if len(fil_time) != 0:
            cod_sen=""
            if two_sensores==True:
                cod_sen=list_est[0][-10:-8]
            
            #-----GRAFICA CONTEO DE GAPS, OVERLAPS y offset
            #-----GRAFICA OFFSET
            
            #col1,col2= st.columns([5,1.1])
            #with col1:
            fig = plt.figure(figsize=(15 , 8))
            ax = fig.add_subplot(312)
            #cuando una estacion entra genera un offset muy grande ese dia, si por ejemplo viene un offet promedio de 2000, pero el dia que entra             nuevamente la estacion ese offset es de 200000, entonces se ve plana la grafica de offset y un pico en el 200000,Para mejorar la             visualizacion, se pone un offset cero el dia que se reconoce que entra la estacion.


            ###### Esto es para que el valor de offset despues de l 0 sea 0 tambien
            busca_ceros=[]
            i=0
            
            for e in fil_time["offs"]:
                #print("##file_time offset |",e)
                if e == 0:
                    busca_ceros.append(i)
                if len(busca_ceros)>0:
                    if e!=0:
                        
                        fil_time["offs"].iloc[i]=0
                        busca_ceros=[]
                if e == -9:
                    #fil_time["offs"].iloc[i]=None
                    fil_time.loc[i,"offs"]=None
                #print("#cada offset", fil_time["offs"].iloc[i])
                

                i=i+1

            ######
            

            lista_offset_sin_ceros=[]
            for of in fil_time["offs"]:
                if of != 0 and str(of) != "nan":
                    
                    lista_offset_sin_ceros.append(of)
                    
            
            
            if len(lista_offset_sin_ceros) >0:
                offset_prom=round(np.mean(np.array(lista_offset_sin_ceros)),2) #Promedio del offset
                max_offset=round(max(lista_offset_sin_ceros),2)
                min_offset=round(min(lista_offset_sin_ceros),2)
            else:
                offset_prom=0   
                max_offset=""
                min_offset=""

            #------para mejorar visalizacion offset. Los ceros puestos que son cuando no se pudo calcular (por no disponibilidad)
            #los convierte en el promedio. Si existen valores gigantes, 50 veces mayor a la media grafica el offset en escala log
            escala="linear"
            i=0       
            for e in fil_time["offs"]:
                if e==0:
                    fil_time["offs"].iloc[i]=offset_prom
                if abs(e) >abs(50*offset_prom) or abs(e)>100000:
                    escala="log"
                i=i+1
            #--------  
            
                    
                    
            if escala == "linear":
                ax.plot(fil_time["fecha"],fil_time["offs"],linewidth=1,color="#ff0000",label="Offset")   
                ax.hlines(offset_prom, fecha_ini, fecha_fin, colors='#000000', linestyles='--', linewidth=0.6)
                
                ax.text(fil_time["fecha"].iloc[int(len(fil_time["fecha"])/2)], offset_prom, "media: "+str(offset_prom), verticalalignment='top',fontsize=9)
                ax.set_ylabel('Offset', color='k', fontsize=14)

            if escala == "log": #si esta en escala lineal obtiene el abs del offset (para poder graficar valores negativos del offset original, no                     grafica media porque pierde sentido, esto sucede porque hay valores gigantes respecto a la media)
                ax.plot(fil_time["fecha"],np.abs(np.array(list(fil_time["offs"]))),linewidth=1,color="#ff0000",label="Offset")         
                ax.set_ylabel('abs (Offset)', color='k', fontsize=14)

            ax.legend(loc='upper left')        
            ax.xaxis.grid(True, which='major')
            ax.yaxis.grid(True, which='major')
            ax.xaxis.set_tick_params(labelsize=9)
            ax.set_yscale(escala)
            #ax.set_xlim( fecha_ini-timedelta(days=1), fecha_fin+timedelta(days=dias_ad))

            #overlaps

            #ax1 = ax.twinx()
            #lista_overlaps_sin_nueves=[]
            #for i in range(0,len(fil_time["fecha"])):
                
            #    if fil_time["num_overlaps"].iloc[i] != -9: #se le puso -9 a los datos no disponibles
            #        lista_overlaps_sin_nueves.append(fil_time["num_overlaps"].iloc[i])
            #        ax1.vlines(fil_time["fecha"].iloc[i]-timedelta(minutes=5), 0, fil_time["num_overlaps"].iloc[i], colors='#32cd32', linewidth=1.5,alpha=0.8)
            #print(lista_overlaps_sin_nueves)
            #ax1.set_ylabel('Conteo de \n(Overlaps, Gaps)', color='k', fontsize=14)
            #ax1.xaxis.grid(True, which='major')
            #ax1.xaxis.set_tick_params(labelsize=9)
            #ax1.vlines(fecha_ini, 0, 0, colors='#32cd32', linewidth=1.5, label='Overlaps')   

            #overlaps_prom=round(np.mean(np.array(lista_overlaps_sin_nueves)),2)   #Promedio overlaps
            #num_overlaps = np.sum(lista_overlaps_sin_nueves)
            #max_overlap=max(lista_overlaps_sin_nueves)

            # gaps
            #lista_gaps_sin_nueves=[]
            #for i in range(0,len(fil_time["fecha"])):
            #    if fil_time["num_gaps"].iloc[i] != -9: #se le puso -9 a los datos no disponibles
            #        lista_gaps_sin_nueves.append(fil_time["num_gaps"].iloc[i])
            #        ax1.vlines(fil_time["fecha"].iloc[i]+timedelta(minutes=5), 0, fil_time["num_gaps"].iloc[i], colors='k', linewidth=1.5,alpha=0.8)

            #ax1.vlines(fecha_fin, 0, 0, colors='k', linewidth=1.5, label='Gaps')            
            #ax1.legend(loc='upper right')
            plt.title('\n\n\nOffset ' + estacion + " "+sensor+can+" | "+fecha_ini.strftime("%Y-%m-%d")+" - "+fecha_fin.strftime("%Y-%m-%d"))

            #gaps_prom = round(np.mean(np.array(lista_gaps_sin_nueves)),2)
            #num_gaps = np.sum(lista_gaps_sin_nueves)
            #max_gaps = max(lista_gaps_sin_nueves)
            
            plt.savefig(folder+f"offset_{estacion}_{sensor}{can}{cod_sen}.png", bbox_inches='tight')
            #st.pyplot(fig=fig, )
            
            #with col2:
                        
            #st.write(f"Max. offset : {max_offset}" )
            #st.markdown(f"**:red[Min. offset:]** {min_offset} **:red[| Max. offset :]** {max_offset}" )
            #st.markdown(f"**:red[Promedio offset:]** {offset_prom}" )
            
            #st.write(f"Max. gaps: {max_gaps}")
            #st.markdown(f"**Num. gaps:** {num_gaps} **| Max. gaps:** {max_gaps}")
            #st.markdown(f"**Promedio de gaps:** {gaps_prom}")
            
            #st.write(f"Max. overlaps: {max_overlap}")
            #st.markdown(f"**:green[Num. overlaps:]** {num_overlaps} **:green[| Max. overlaps:]** {max_overlap}")
            #st.markdown(f"**:green[Promedio overlaps:]** {overlaps_prom}")
        else:
            print(f"No hay datos para la estación {estacion} entre las fechas del {fecha_ini} {fecha_fin}")    
               
    def plot_gapover_fij(self, can, folder,two_sensores=False):
            
            #folder = self.folder
            estacion = self.estacion
            sensor = self.sensor
            fecha_ini = self.fecha_ini
            fecha_fin = self.fecha_fin
            fil_time = self.fil_time
            dias_ad = self.dias_ad  
            list_est= self.list_est 
            
            if len(fil_time) != 0:
                cod_sen=""
                if two_sensores==True:
                    cod_sen=list_est[0][-10:-8]
                
                
                #-----GRAFICA CONTEO DE GAPS, OVERLAPS y offset
                #-----GRAFICA OFFSET
                
                #col1,col2= st.columns([5,1.1])
                #with col1:
                fig = plt.figure(figsize=(15 , 8))
                ax = fig.add_subplot(312)
                #cuando una estacion entra genera un offset muy grande ese dia, si por ejemplo viene un offet promedio de 2000, pero el dia que entra             nuevamente la estacion ese offset es de 200000, entonces se ve plana la grafica de offset y un pico en el 200000,Para mejorar la             visualizacion, se pone un offset cero el dia que se reconoce que entra la estacion.


                ###### Esto es para que el valor de offset despues de l 0 sea 0 tambien
                busca_ceros=[]
                i=0
                
                #for e in fil_time["offs"]:
                    #print("##file_time offset |",e)
                #    if e == 0:
                #        busca_ceros.append(i)
                #    if len(busca_ceros)>0:
                #        if e!=0:
                            
                #            fil_time["offs"].iloc[i]=0
                #            busca_ceros=[]
                #    if e == -9:
                #        fil_time["offs"].iloc[i]=None
                    #print("#cada offset", fil_time["offs"].iloc[i])
                    

                #    i=i+1

                ######
                

                #lista_offset_sin_ceros=[]
                #for of in fil_time["offs"]:
                #    if of != 0 and str(of) != "nan":
                        
                #        lista_offset_sin_ceros.append(of)
                        
                
                
                #if len(lista_offset_sin_ceros) >0:
                #    offset_prom=round(np.mean(np.array(lista_offset_sin_ceros)),2) #Promedio del offset
                #    max_offset=round(max(lista_offset_sin_ceros),2)
                #    min_offset=round(min(lista_offset_sin_ceros),2)
                #else:
                #    offset_prom=0   
                #    max_offset=""
                #    min_offset=""

                #------para mejorar visalizacion offset. Los ceros puestos que son cuando no se pudo calcular (por no disponibilidad)
                #los convierte en el promedio. Si existen valores gigantes, 50 veces mayor a la media grafica el offset en escala log
                #escala="linear"
                #i=0       
                #for e in fil_time["offs"]:
                #    if e==0:
                #        fil_time["offs"].iloc[i]=offset_prom
                #    if abs(e) >abs(50*offset_prom) or abs(e)>100000:
                #        escala="log"
                #    i=i+1
                #--------  
                
                        
                        
                #if escala == "linear":
                #    ax.plot(fil_time["fecha"],fil_time["offs"],linewidth=1,color="#ff0000",label="Offset")   
                #    ax.hlines(offset_prom, fecha_ini, fecha_fin, colors='#000000', linestyles='--', linewidth=0.6)
                    
                #    ax.text(fil_time["fecha"].iloc[int(len(fil_time["fecha"])/2)], offset_prom, "media: "+str(offset_prom), verticalalignment='top',fontsize=9)
                #    ax.set_ylabel('Offset', color='k', fontsize=14)

                #if escala == "log": #si esta en escala lineal obtiene el abs del offset (para poder graficar valores negativos del offset original, no                     grafica media porque pierde sentido, esto sucede porque hay valores gigantes respecto a la media)
                #    ax.plot(fil_time["fecha"],np.abs(np.array(list(fil_time["offs"]))),linewidth=1,color="#ff0000",label="Offset")         
                #    ax.set_ylabel('abs (Offset)', color='k', fontsize=14)

                ax.plot(fecha_ini,0)
                ax.legend(loc='upper left')        
                ax.xaxis.grid(True, which='major')
                ax.yaxis.grid(True, which='major')
                ax.xaxis.set_tick_params(labelsize=9)
                ax.set_yscale("linear")
                #ax.set_xlim( fecha_ini-timedelta(days=1), fecha_fin+timedelta(days=dias_ad))

                #overlaps

                #ax1 = ax1.twinx()
                lista_overlaps_sin_nueves=[]
                for i in range(0,len(fil_time["fecha"])):
                    
                    if fil_time["num_overlaps"].iloc[i] != -9: #se le puso -9 a los datos no disponibles
                        lista_overlaps_sin_nueves.append(fil_time["num_overlaps"].iloc[i])
                        ax.vlines(fil_time["fecha"].iloc[i]-timedelta(minutes=5), 0, fil_time["num_overlaps"].iloc[i], colors='#32cd32', linewidth=1.5,alpha=0.8)
                #print(lista_overlaps_sin_nueves)
                ax.set_ylabel('Conteo de \n(Overlaps, Gaps)', color='k', fontsize=14)
                ax.xaxis.grid(True, which='major')
                #ax1.yaxis.grid(True, which='major')
                ax.xaxis.set_tick_params(labelsize=9)
                ax.vlines(fecha_ini, 0, 0, colors='#32cd32', linewidth=1.5, label='Overlaps')   

                if len(lista_overlaps_sin_nueves) != 0:
                    overlaps_prom=round(np.mean(np.array(lista_overlaps_sin_nueves)),2)   #Promedio overlaps
                    num_overlaps = np.sum(lista_overlaps_sin_nueves)
                    max_overlap=max(lista_overlaps_sin_nueves)
                if len(lista_overlaps_sin_nueves) == 0:
                    overlaps_prom=0
                    num_overlaps=0
                    max_overlap=0
                # gaps
                lista_gaps_sin_nueves=[]
                for i in range(0,len(fil_time["fecha"])):
                    if fil_time["num_gaps"].iloc[i] != -9: #se le puso -9 a los datos no disponibles
                        lista_gaps_sin_nueves.append(fil_time["num_gaps"].iloc[i])
                        ax.vlines(fil_time["fecha"].iloc[i]+timedelta(minutes=5), 0, fil_time["num_gaps"].iloc[i], colors='k', linewidth=1.5,alpha=0.8)

                ax.vlines(fecha_fin, 0, 0, colors='k', linewidth=1.5, label='Gaps')            
                ax.legend(loc='upper right')
                plt.title('\n\n\nConteo de (Gaps - Overlaps) ' + estacion + " "+sensor+can+" | "+fecha_ini.strftime("%Y-%m-%d")+" - "+fecha_fin.strftime("%Y-%m-%d"))

                if len(lista_gaps_sin_nueves) !=0:
                    
                    gaps_prom = round(np.mean(np.array(lista_gaps_sin_nueves)),2)
                    num_gaps = np.sum(lista_gaps_sin_nueves)
                    max_gaps = max(lista_gaps_sin_nueves)
                if len(lista_gaps_sin_nueves) ==0:
                    gaps_prom = 0
                    num_gaps=0
                    max_gaps=0
                
                plt.savefig(folder+f"gapover_{estacion}_{sensor}{can}{cod_sen}.png", bbox_inches='tight')
                #st.pyplot(fig=fig, )


                
                #with col2:
                            
                #st.write(f"Max. offset : {max_offset}" )
                #st.markdown(f"**:red[Min. offset:]** {min_offset} **:red[| Max. offset :]** {max_offset}" )
                #st.markdown(f"**:red[Promedio offset:]** {offset_prom}" )
                
                #st.write(f"Max. gaps: {max_gaps}")
                #st.markdown(f"**Num. gaps:** {num_gaps} **| Max. gaps:** {max_gaps}")
                #st.markdown(f"**Promedio de gaps:** {gaps_prom}")
                
                #st.write(f"Max. overlaps: {max_overlap}")
                #st.markdown(f"**:green[Num. overlaps:]** {num_overlaps} **:green[| Max. overlaps:]** {max_overlap}")
                #st.markdown(f"**:green[Promedio overlaps:]** {overlaps_prom}")
            else:
                print(f"No hay datos para la estación {estacion} entre las fechas del {fecha_ini} {fecha_fin}")  
    
    def plot_ppsd_fij(self, can, folder,two_sensores=False):
        
        #folder = self.folder
        estacion = self.estacion
        sensor = self.sensor
        fecha_ini = self.fecha_ini
        fecha_fin = self.fecha_fin
        fil_time = self.fil_time
        dias_ad = self.dias_ad  
        list_est= self.list_est 
        
        if len(fil_time) != 0:
            cod_sen=""
            if two_sensores==True:
                cod_sen=list_est[0][-10:-8]
            
            #-------GRFICA DE PPSD, CONTEO PICOS Y PICOS EN EL DÍA
            
            #col1,col2= st.columns([5,1.1])
            #with col1:
            fig = plt.figure(figsize=(15 , 8))
            ax = fig.add_subplot(313)
            
            #ppsd sin -9
            lista_ppsd_sin_menosnueves=[]
            lista_ppsd=[]
            for of in fil_time["p_ppsd"]:
                if of != -9:
                    lista_ppsd_sin_menosnueves.append(of)
                    lista_ppsd.append(of)
                    #print(type(of),of)
                if of == -9:
                    lista_ppsd.append(None)
                
                
            
            #st.dataframe(fil_time["p_ppsd"])
        
            #plotear promedio ppsd
            ax.plot(fil_time["fecha"],lista_ppsd,linewidth=2,color="#0D9000",label="ppsd")   
            #ax.vlines(fecha_fin, list(lista_ppsd)[0], 0, colors='#0D9000', linewidth=1.0, label='PPSD (%)fuera')        
            ax.set_ylabel('PPSD % por fuera', color='k', fontsize=14)                
            ax.legend(loc='upper left')
            ax.xaxis.grid(True, which='major')
            #ax.set_xlim( fecha_ini-timedelta(days=1), fecha_fin+timedelta(days=dias_ad))
            ax.xaxis.set_tick_params(labelsize=9)
            #picos
            ax1 = ax.twinx()
            lista_picos_sin_nueves=[]
            
            for i in range(0,len(fil_time["fecha"])):
                if fil_time["peaks"].iloc[i] != -9 : #se le puso -9 a los datos no disponibles
                    lista_picos_sin_nueves.append(fil_time["peaks"].iloc[i])                
                    ax1.vlines(fil_time["fecha"].iloc[i], 0, fil_time["peaks"].iloc[i], colors='#8000ff', linewidth=1.0,alpha=1,zorder=100)
            
            ax1.vlines(fecha_fin, 0, 0, colors='#8000ff', linewidth=1.0, label='Picos (conteo)')        
            ax1.set_ylabel('Conteo de \nPicos', color='k', fontsize=14)                
            ax1.legend(loc='upper left')    
            plt.title('\n\n\n %ppsd y picos ' + estacion + " "+sensor+can+" | "+fecha_ini.strftime("%Y-%m-%d")+" - "+fecha_fin.strftime("%Y-%m-%d"))
                    
            ax1.yaxis.grid(True, which='major')   
            
            if len(lista_ppsd_sin_menosnueves) != 0: 
                ppsd_prom = round(np.mean(np.array(lista_ppsd_sin_menosnueves)),2)
            if len(lista_ppsd_sin_menosnueves) == 0: 
                ppsd_prom = None
            if len(lista_picos_sin_nueves) != 0:
                num_picos = np.sum(lista_picos_sin_nueves)
                max_picos = max(lista_picos_sin_nueves)
            if len(lista_picos_sin_nueves) == 0:
                num_picos =0
                max_picos =0
            
            #EN TIEMPO
            """
            l_fil_tpics = []
            for fil_time_pic in fil_time["time_peaks"]:
                if len(str(fil_time_pic)) > 2 :
                    l_fil_tpics.append(fil_time_pic)

            ax2 = ax.twinx()
            if len(l_fil_tpics) > 0:
                
                to_r=datetime(1,1,1,0,0,1)
                rectangulo3 = patches.Rectangle( ( fecha_fin, to_r ),timedelta(dias_ad),timedelta(hours=24),fill=True,facecolor='b',edgecolor='b',linewidth=0.2,alpha=0.8)
                ax2.add_patch(rectangulo3)

                
                for tp in fil_time["time_peaks"]:
                    tdia=datetime(tp.year,tp.month,tp.day,12,0,0)
                    to=datetime(1,1,1,tp.hour,tp.minute,tp.second)                    
                    ax1.scatter(tdia,to,s=1.5,c="k",alpha=0.5,zorder=100)
                    ax1.hlines(to, tdia, max(lista_tiempo), colors='k', linestyles='--', linewidth=0.1,alpha=0.7) 
                    ax1.hlines(to, lista_tiempo[-1], lista_tiempo[-1]+timedelta(dias_ad), colors='k', linestyles='-', linewidth=0.2,alpha=0.7)
            """
            
            plt.savefig(folder+f"ppsd_{estacion}_{sensor}{can}{cod_sen}.png", bbox_inches='tight')
            #st.pyplot(fig=fig, )
                
            #with col2:
                
                
            #st.write("_________________________")    
            #st.markdown(f"**:violet[Num. picos:]** {num_picos} **:violet[| Max. de picos:]** {max_picos}")
            #st.markdown(f"**:green[Promedio (%) ppsd por fuera:]** {ppsd_prom}%")
            #st.write("_________________________")
        
        else:
            print(f"No hay datos para la estación {estacion} entre las fechas del {fecha_ini} {fecha_fin}")  
                
    def plot_disp_din(self, can, save=False, folder=" "):
        
        #folder = self.folder
        estacion = self.estacion
        sensor = self.sensor
        fecha_ini = self.fecha_ini
        fecha_fin = self.fecha_fin
        list_est = self.list_est
        fil_time = self.fil_time
        dias_ad = self.dias_ad 
        
        codigo_l = list_est[0][-10:-8]
        
        if estacion == "ROSC" and sensor=="BH":
            codigo_l = "--"
        
        
        
        
        if len(fil_time) != 0:
        
            
            col1,col2= st.columns([5,1.1])
            with col1:
                fig = make_subplots(specs=[[{"secondary_y": True}]])
                f_i = datetime(fecha_ini.year,fecha_ini.month,fecha_ini.day)

                
                
                #Graficar Disponibilidad
                fig.add_trace(
                    go.Scatter(x=fil_time["fecha"], y=fil_time["disponibilidad"],mode = "lines",marker_color='rgba(0,99,182,10)',connectgaps=False,name='Disponibilidad',
                            hovertemplate="Fecha: %{x}<br>" +"Disponibilidad: %{y}%<br>" +"<extra></extra>",  #hovertemplate, para seleccionar el texto al mostrar al pasar el cursor
                            hoverlabel=dict(bgcolor="#0091E9",font_size=15,font_color="white",font_family="Times New Roman")),secondary_y=False,) #hoverlabel para configurar la forma del texto de información
                fig.add_trace(go.Scatter( x=[f_i], y=[float(list(fil_time["disponibilidad"])[0])-1], mode = "lines",marker_color='rgba(59,59,59,0.7)',name="Gaps (time)",line=dict(width=9),), secondary_y=False)
        
                
                #con este foorloop se selecciona en una lista solo los tiempos de gaps existentes
                l_fil_tgap = []
                for fil_time_gap in fil_time["times_gaps_graf_din"]:
                    if len(str(fil_time_gap[0][0])) > 2:
                        l_fil_tgap.append(fil_time_gap)
                
                
                #para graficar los gaps existentes
                if len(l_fil_tgap) > 0:
                    for tg in l_fil_tgap:
                        
                        for g in tg:
                            
                            #Esta condicional es para que cuando el gap  hace que la fecha final sea el dia siguiente a las 00:00:00, lo cambie por el mismo día a las 23:59:59
                            if g[1].day == g[0].day:
                                dia = [datetime(g[0].year,g[0].month,g[0].day,0,0,0),datetime(g[1].year,g[1].month,g[1].day,0,0,0)]
                                
                                t_gap = [datetime(1,1,1,g[0].hour,g[0].minute,g[0].second,g[0].microsecond), datetime(1,1,1,g[1].hour,g[1].minute,g[1].second,g[1].microsecond)]    
                                fig.add_trace(
                                    go.Scatter(x=dia, y=t_gap,mode='lines', connectgaps=False,marker_color='rgba(59,59,59,0.7)',name='Gaps (time)',showlegend=False,line=dict(width=10),
                                            hovertemplate="Fecha: %{x}<br>" +"Gap: %{y}<br>" +"<extra></extra>",#hovertemplate, para seleccionar el texto al mostrar al pasar el cursor
                                            hoverlabel=dict(bgcolor="#A0A0A0",font_size=15,font_color="white",font_family="Times New Roman")),  secondary_y=True,)#hoverlabel para configurar la forma del texto de información
                                
                            else:
                                if g[1].month == g[0].month:
                                    dia = [datetime(g[0].year,g[0].month,g[0].day,0,0,0),datetime(g[1].year,g[1].month,g[0].day,0,0,0)]
                                    
                                    t_gap = [datetime(1,1,1,g[0].hour,g[0].minute,g[0].second,g[0].microsecond), datetime(1,1,1,23,59,59)]    
                                    fig.add_trace(
                                        go.Scatter(x=dia, y=t_gap,mode='lines', connectgaps=False,marker_color='rgba(59,59,59,0.7)',name='Gaps (time)',showlegend=False,line=dict(width=10),
                                                hovertemplate="Fecha: %{x}<br>" +"Gap: %{y}<br>" +"<extra></extra>",#hovertemplate, para seleccionar el texto al mostrar al pasar el cursor
                                                hoverlabel=dict(bgcolor="#A0A0A0",font_size=15,font_color="white",font_family="Times New Roman")),  secondary_y=True,)#hoverlabel para configurar la forma del texto de información
                                else:
                                    dia = [datetime(g[0].year,g[0].month,g[0].day,0,0,0),datetime(g[1].year,g[0].month,g[0].day,0,0,0)]
                                    
                                    t_gap = [datetime(1,1,1,g[0].hour,g[0].minute,g[0].second,g[0].microsecond), datetime(1,1,1,23,59,59)]    
                                    fig.add_trace(
                                        go.Scatter(x=dia, y=t_gap,mode='lines', connectgaps=False,marker_color='rgba(59,59,59,0.7)',name='Gaps (time)',showlegend=False,line=dict(width=10),
                                                hovertemplate="Fecha: %{x}<br>" +"Gap: %{y}<br>" +"<extra></extra>",#hovertemplate, para seleccionar el texto al mostrar al pasar el cursor
                                                hoverlabel=dict(bgcolor="#A0A0A0",font_size=15,font_color="white",font_family="Times New Roman")),  secondary_y=True,)#hoverlabel para configurar la forma del texto de información

                
                            
                fig.update_yaxes(range=(datetime(1,1,1,0,0,0),datetime(1,1,1,23,59,59)), tickformat="%H:%M:%S\nHora",)  #Limites en eje y horas
                
                #detalles
                tit = f"Disponibilidad {estacion}  {sensor}{can} |  {fecha_ini.strftime('%Y-%m-%d')} - {fecha_fin.strftime('%Y-%m-%d')}"
                
                fig.update_layout(width = 800, height = 310,title_text=tit, 
                                legend=dict(orientation="h",yanchor="bottom",y=1.1,xanchor="right",x=1),
                                yaxis_title='Disponibilidad (%)')   
                
                fig.update_xaxes(showspikes=True,spikethickness=1,spikesnap="cursor",spikemode="across")#este es para colocar las gias al pasar el cursor
                fig.update_yaxes(showspikes=True,spikethickness=1,spikesnap="cursor",spikemode="across")#este es para colocar las gias al pasar el cursor
                
                fig.update_xaxes(showgrid=True, ticklabelmode="period",tickformat="%d\n%b\n%Y", gridwidth=1, gridcolor='LightGrey')
                fig.update_xaxes(showline=True, linewidth=2, linecolor='lightgray', mirror=True)
                fig.update_yaxes(showline=True, linewidth=2, linecolor='lightgray', mirror=True, )
                
                if save==True:
                    #fig.write_image(folder+f"disp_{estacion}_{sensor}{can}.png")
                    #plotly.io.write_image(fig, folder+f"disp_{estacion}_{sensor}{can}.png")
                    imag = fig.to_image(format="png")
                    with open(folder+f"disp_{estacion}_{sensor}{can}.png"   , "wb") as f:
                        f.write(imag)
                
                st.plotly_chart(fig, use_container_width=True)
            
            with col2:
                
                fil_time2 = fil_time[(fil_time["disponibilidad"] != -9)]
                min_dis = fil_time2["disponibilidad"].min()             #minima desponibilidad
                max_dis = fil_time2["disponibilidad"].max()             #maxima disponibilidad
                prom_dis = fil_time2["disponibilidad"].mean()           #promedio disponibilidad
                
                st.write(".")
                st.markdown(f"**:blue[Min. disp.:]** {min_dis}%**:blue[| Max. disp.:]** {max_dis}%")
                #st.write(f"Max. disp: {max_dis}")
                st.markdown(f"**:blue[Promedio disp.:]** {round(prom_dis,2)}%")
                st.write("_________________________")
            
            return [min_dis,max_dis,prom_dis,codigo_l]
        
        else:
            st.error("No hay datos de esta estación en estas fechas")
            return [0,0,0,codigo_l]
    
    def plot_off_din(self, can):

        #folder = self.folder
        estacion = self.estacion
        sensor = self.sensor
        fecha_ini = self.fecha_ini
        fecha_fin = self.fecha_fin
        list_est = self.list_est
        fil_time = self.fil_time
        dias_ad = self.dias_ad 
        
        #-----GRAFICA CONTEO DE GAPS, OVERLAPS y offset
        #-----GRAFICA OFFSET
        if len(fil_time) != 0:
            col1,col2= st.columns([5,1.1])
            with col1:
                ###### Esto es para que el valor de offset despues de l 0 sea 0 tambien
                #busca_ceros=[]
                #i=0
                
                #for e in fil_time["offs"]:
                    
                    #if e == 0:
                    #    busca_ceros.append(i)
                    #if len(busca_ceros)>0:
                    #    if e!=0:
                            
                    #        fil_time["offs"].iloc[i]=0
                    #        busca_ceros=[]
                    #if e == -9:
                        #fil_time["offs"].iloc[i]=None
                    #    fil_time.loc[i,"offs"]=None
                    
                    #i=i+1
                
                ###Datos promedio, y mas
                lista_offset_sin_ceros=[]
                for of in fil_time["offs"]:
                    if str(of) != "nan" and of != None and of != -9 :
                        
                        lista_offset_sin_ceros.append(of)
                        
                
                
                if len(lista_offset_sin_ceros) >0:
                    offset_prom=round(np.mean(np.array(lista_offset_sin_ceros)),2) #Promedio del offset
                    max_offset=round(max(lista_offset_sin_ceros),2)
                    min_offset=round(min(lista_offset_sin_ceros),2)
                else:
                    offset_prom=0   
                    max_offset=""
                    min_offset=""
                ######
                
                #------para mejorar visalizacion offset. Los ceros puestos que son cuando no se pudo calcular (por no disponibilidad)
                #los convierte en el promedio. Si existen valores gigantes, 50 veces mayor a la media grafica el offset en escala log
                escala="linear"
                i=0       
                for e in fil_time["offs"]:
                    if e==0:
                        fil_time["offs"].iloc[i]=offset_prom
                    if abs(e) >abs(50*offset_prom) or abs(e)>100000:
                        #escala="log"
                        escala="linear"
                    i=i+1
                #-------- 
                
                l2_off =[]
                for f in fil_time["offs"]:
                    if f == 0 or f == -9:
                        l2_off.append(None)
                    else:
                        l2_off.append(f)
                fil_time["offs"] = l2_off
                
                fig2 = make_subplots(specs=[[{"secondary_y": True}]])
                
                
                if escala == "linear":
                    fig2.add_trace(
                        go.Scatter( x=fil_time["fecha"], y=fil_time["offs"], mode = "lines",marker_color='rgba(225,0,0,1)', connectgaps=False,name='offset',
                                hovertemplate="Fecha: %{x}<br>" +"Offset: %{y}<br>" +"<extra></extra>",  #hovertemplate, para seleccionar el texto al mostrar al pasar el cursor
                                hoverlabel=dict(bgcolor="#C11515",font_size=15,font_color="white",font_family="Times New Roman")),secondary_y=False,) #hoverlabel para configurar la forma del texto de información
                    fig2.add_hline(y=offset_prom,line_dash="dot",line_width=1,line_color="grey",annotation_text=f"media: {offset_prom}", 
                    annotation_position="bottom")
                    #fig2.update_yaxes(type="log")

                if escala == "log": #si esta en escala lineal obtiene el abs del offset (para poder graficar valores negativos del offset original, no                     grafica media porque pierde sentido, esto sucede porque hay valores gigantes respecto a la media)
                    fig2.add_trace(go.Scatter( x=fil_time["fecha"], y=np.abs(np.array(list(fil_time["offs"]))), mode = "lines",marker_color='rgba(225,0,0,1)', connectgaps=True,name='offset_abs',),secondary_y=False,)
                    fig2.add_trace(go.Scatter( x=[fecha_ini], y=[0], mode = "lines",marker_color='rgba(140,140,140,1)',name="Gaps",line=dict(width=9),), secondary_y=False)
                
                #leyenda overlaps
                fig2.add_trace(
                    go.Scatter( x=[fecha_ini ], y=[list(fil_time["offs"])[0]], mode = "lines",marker_color='rgba(50,205,50,0.7)',name="Overlaps",line=dict(width=2),),secondary_y=True,) #hoverlabel para configurar la forma del texto de información
                #leyenda de gaps
                fig2.add_trace(
                    go.Scatter( x=[fecha_ini ], y=[list(fil_time["offs"])[0]], mode = "lines",marker_color='rgba(71,71,71,0.7)',name="Gaps",line=dict(width=2),),secondary_y=True,) #hoverlabel para configurar la forma del texto de información

                #para graficar conteo de overlaps
                for n_over,n_gaps,time in zip(fil_time["num_overlaps"],fil_time["num_gaps"],fil_time["fecha"]):
                    
                    #graficar numero de overlaps por dia
                    fig2.add_trace(go.Scatter(x=[time,time], y=[0,n_over],text=[0,n_gaps],mode='lines', connectgaps=False,marker_color='rgba(50,205,50,0.7)',name='Overlaps',showlegend=False   ,line=dict(width=2),
                                hovertemplate="Fecha: %{x}<br>" +"Overlaps: %{y}<br>"+"Gaps: %{text}<br>" +"<extra></extra>",  #hovertemplate, para seleccionar el texto al mostrar al pasar el cursor
                                hoverlabel=dict(bgcolor="#34C115",font_size=15,font_color="white",font_family="Times New Roman")),secondary_y=True,) #hoverlabel para configurar la forma del texto de información
                    #graficar numero de gaps por dia
                    fig2.add_trace(go.Scatter(x=[time,time], y=[0,n_gaps],text=[0,n_over],mode='lines', connectgaps=False,marker_color='rgba(71,71,71,0.7)',name='Gaps',showlegend=False   ,line=dict(width=2),
                                hovertemplate="Fecha: %{x}<br>" +"Gaps: %{y}<br>"+"Overlaps: %{text}<br>" +"<extra></extra>",  #hovertemplate, para seleccionar el texto al mostrar al pasar el cursor
                            hoverlabel=dict(bgcolor="#797A79",font_size=15,font_color="white",font_family="Times New Roman")),secondary_y=True,) #hoverlabel para configurar la forma del texto de información
                
                ##Datos promedio y mas
                lista_overlaps_sin_nueves=[]
                for i in range(0,len(fil_time["fecha"])):
                    if np.isnan(fil_time["num_overlaps"].iloc[i]):
                        continue
                    if fil_time["num_overlaps"].iloc[i] != -9: #se le puso -9 a los datos no disponibles
                        lista_overlaps_sin_nueves.append(fil_time["num_overlaps"].iloc[i])
                
                if len(lista_overlaps_sin_nueves) != 0:     
                    overlaps_prom=round(np.mean(np.array(lista_overlaps_sin_nueves)),2)   #Promedio overlaps
                    num_overlaps = np.sum(lista_overlaps_sin_nueves)
                    max_overlap=max(lista_overlaps_sin_nueves)
                if len(lista_overlaps_sin_nueves) == 0:
                    overlaps_prom=0
                    num_overlaps=0
                    max_overlap=0 
                
                # promedio de gaps y mas
                lista_gaps_sin_nueves=[]
                for i in range(0,len(fil_time["fecha"])):
                    if np.isnan(fil_time["num_gaps"].iloc[i]):
                        continue
                    if fil_time["num_gaps"].iloc[i] != -9: #se le puso -9 a los datos no disponibles
                        lista_gaps_sin_nueves.append(fil_time["num_gaps"].iloc[i])

                if len(lista_gaps_sin_nueves)!=0:
                    gaps_prom = round(np.mean(np.array(lista_gaps_sin_nueves)),2)
                    num_gaps = np.sum(lista_gaps_sin_nueves)
                    max_gaps = max(lista_gaps_sin_nueves)
                if len(lista_gaps_sin_nueves)==0:
                    gaps_prom=0
                    num_gaps=0
                    max_gaps=0
                #limites eje auxiliar o secundario
                fig2.update_yaxes(range=(0,int(max_overlap)),secondary_y=True) #Limites en eje conteo
                fig2.update_yaxes(range=(0,int(max_gaps)),title_text="Conteo (Overlaps, Gaps)",secondary_y=True)  #Limites en eje conteo

                #detalles
                tit2 = 'Conteo de (Gaps - Overlaps) y Offset ' + estacion + " "+sensor+can+" | "+fecha_ini.strftime("%Y-%m-%d")+" - "+fecha_fin.strftime("%Y-%m-%d")
                fig2.update_layout(width = 800, height = 310,title_text=tit2,
                                legend=dict(orientation="h",yanchor="bottom",y=1.1,xanchor="right",x=1),
                                yaxis_title='Offset')
                
                fig2.update_xaxes(showspikes=True,spikethickness=1,spikesnap="cursor",spikemode="across")#este es para colocar las gias al pasar el cursor
                fig2.update_yaxes(showspikes=True,spikethickness=1,spikesnap="cursor",spikemode="across")#este es para colocar las gias al pasar el cursor
                
                fig2.update_xaxes(showgrid=True, ticklabelmode="period",tickformat="%d\n%b\n%Y", gridwidth=1, gridcolor='LightGrey')
                fig2.update_xaxes(showline=True, linewidth=2, linecolor='lightgray', mirror=True)
                fig2.update_yaxes(showline=True, linewidth=2, linecolor='lightgray', mirror=True, )
                
                st.plotly_chart(fig2, use_container_width=True)

            with col2:
                        
                #st.write(f"Max. offset : {max_offset}" )
                
                st.write(".")
                st.markdown(f"**:red[Min. offset:]** {min_offset} **:red[| Max. offset :]** {max_offset}" )
                st.markdown(f"**:red[Promedio offset:]** {offset_prom}" )
                
                #st.write(f"Max. gaps: {max_gaps}")
                st.markdown(f"**Num. gaps:** {num_gaps} **| Max. gaps:** {max_gaps}")
                st.markdown(f"**Promedio de gaps:** {gaps_prom}")
                
                #st.write(f"Max. overlaps: {max_overlap}")
                st.markdown(f"**:green[Num. overlaps:]** {num_overlaps} **:green[| Max. overlaps:]** {max_overlap}")
                st.markdown(f"**:green[Promedio overlaps:]** {overlaps_prom}")
        
        else:
            st.error("No hay datos de esta estación en estas fechas")

    def plot_off2_din(self, can):

            #folder = self.folder
            estacion = self.estacion
            sensor = self.sensor
            fecha_ini = self.fecha_ini
            fecha_fin = self.fecha_fin
            list_est = self.list_est
            fil_time = self.fil_time
            dias_ad = self.dias_ad 
            
            codigo_l = list_est[0][-10:-8]
            if len(fil_time) != 0:
                
                
                #-----GRAFICA CONTEO DE GAPS, OVERLAPS y offset
                #-----GRAFICA OFFSET
                
                col1,col2= st.columns([5,1.1])
                with col1:
                    ###### Esto es para que el valor de offset despues de l 0 sea 0 tambien
                    busca_ceros=[]
                    i=0
                    
                    for e in fil_time["offs"]:
                        
                        if e == 0:
                            busca_ceros.append(i)
                        if len(busca_ceros)>0:
                            if e!=0:
                                
                                fil_time["offs"].iloc[i]=0
                                busca_ceros=[]
                        if e == -9:
                                #fil_time["offs"].iloc[i]=None
                                fil_time.loc[i,"offs"]=None
                        
                        i=i+1
                    
                    ###Datos promedio, y mas
                    lista_offset_sin_ceros=[]
                    for of in fil_time["offs"]:
                        if of != 0 and str(of) != "nan":
                            
                            lista_offset_sin_ceros.append(of)
                            
                    
                    
                    if len(lista_offset_sin_ceros) >0:
                        offset_prom=round(np.mean(np.array(lista_offset_sin_ceros)),2) #Promedio del offset
                        max_offset=round(max(lista_offset_sin_ceros),2)
                        min_offset=round(min(lista_offset_sin_ceros),2)
                    else:
                        offset_prom=0   
                        max_offset=""
                        min_offset=""
                    ######
                    
                    #------para mejorar visalizacion offset. Los ceros puestos que son cuando no se pudo calcular (por no disponibilidad)
                    #los convierte en el promedio. Si existen valores gigantes, 50 veces mayor a la media grafica el offset en escala log
                    escala="linear"
                    i=0       
                    for e in fil_time["offs"]:
                        if e==0:
                            fil_time["offs"].iloc[i]=offset_prom
                        if e==-9:
                            fil_time["offs"].iloc[i]=None
                        if abs(e) >abs(50*offset_prom) or abs(e)>100000:
                            #escala="log"
                            escala="linear"
                        i=i+1
                    #-------- 
                    
                    fig2 = make_subplots(specs=[[{"secondary_y": True}]])
                    
                    
                    if escala == "linear":
                        fig2.add_trace(
                            go.Scatter( x=fil_time["fecha"], y=fil_time["offs"], mode = "lines",marker_color='rgba(225,0,0,1)', connectgaps=False,name='offset',
                                    hovertemplate="Fecha: %{x}<br>" +"Offset: %{y}<br>" +"<extra></extra>",  #hovertemplate, para seleccionar el texto al mostrar al pasar el cursor
                                    hoverlabel=dict(bgcolor="#C11515",font_size=15,font_color="white",font_family="Times New Roman")),secondary_y=False,) #hoverlabel para configurar la forma del texto de información
                        fig2.add_hline(y=offset_prom,line_dash="dot",line_width=1,line_color="grey",annotation_text=f"media: {offset_prom}", 
                        annotation_position="bottom")
                        #fig2.update_yaxes(type="log")

                    if escala == "log": #si esta en escala lineal obtiene el abs del offset (para poder graficar valores negativos del offset original, no                     grafica media porque pierde sentido, esto sucede porque hay valores gigantes respecto a la media)
                        fig2.add_trace(go.Scatter( x=fil_time["fecha"], y=np.abs(np.array(list(fil_time["offs"]))), mode = "lines",marker_color='rgba(225,0,0,1)', connectgaps=True,name='offset_abs',),secondary_y=False,)
                        fig2.add_trace(go.Scatter( x=[f_i], y=[0], mode = "lines",marker_color='rgba(140,140,140,1)',name="Gaps",line=dict(width=9),), secondary_y=False)
                    
                    #leyenda overlaps
                    #fig2.add_trace(
                    #    go.Scatter( x=[fecha_ini ], y=[0], mode = "lines",marker_color='rgba(50,205,50,0.7)',name="Overlaps",line=dict(width=2),),secondary_y=True,) #hoverlabel para configurar la forma del texto de información
                    #leyenda de gaps
                    #fig2.add_trace(
                    #    go.Scatter( x=[fecha_ini ], y=[0], mode = "lines",marker_color='rgba(71,71,71,0.7)',name="Gaps",line=dict(width=2),),secondary_y=True,) #hoverlabel para configurar la forma del texto de información

                    #para graficar conteo de overlaps
                    #for n_over,n_gaps,time in zip(fil_time["num_overlaps"],fil_time["num_gaps"],fil_time["fecha"]):
                        
                        #graficar numero de overlaps por dia
                    #    fig2.add_trace(go.Scatter(x=[time,time], y=[0,n_over],text=[0,n_gaps],mode='lines', connectgaps=False,marker_color='rgba(50,205,50,0.7)',name='Overlaps',showlegend=False   ,line=dict(width=2),
                    #                hovertemplate="Fecha: %{x}<br>" +"Overlaps: %{y}<br>"+"Gaps: %{text}<br>" +"<extra></extra>",  #hovertemplate, para seleccionar el texto al mostrar al pasar el cursor
                    #                hoverlabel=dict(bgcolor="#34C115",font_size=15,font_color="white",font_family="Times New Roman")),secondary_y=True,) #hoverlabel para configurar la forma del texto de información
                        #graficar numero de gaps por dia
                    #    fig2.add_trace(go.Scatter(x=[time,time], y=[0,n_gaps],text=[0,n_over],mode='lines', connectgaps=False,marker_color='rgba(71,71,71,0.7)',name='Gaps',showlegend=False   ,line=dict(width=2),
                    #                hovertemplate="Fecha: %{x}<br>" +"Gaps: %{y}<br>"+"Overlaps: %{text}<br>" +"<extra></extra>",  #hovertemplate, para seleccionar el texto al mostrar al pasar el cursor
                    #            hoverlabel=dict(bgcolor="#797A79",font_size=15,font_color="white",font_family="Times New Roman")),secondary_y=True,) #hoverlabel para configurar la forma del texto de información
                    
                    ##Datos promedio y mas
                    #lista_overlaps_sin_nueves=[]
                    #for i in range(0,len(fil_time["fecha"])):
                        
                    #    if fil_time["num_overlaps"].iloc[i] != -9: #se le puso -9 a los datos no disponibles
                    #        lista_overlaps_sin_nueves.append(fil_time["num_overlaps"].iloc[i])
                            
                    #overlaps_prom=round(np.mean(np.array(lista_overlaps_sin_nueves)),2)   #Promedio overlaps
                    #num_overlaps = np.sum(lista_overlaps_sin_nueves)
                    #max_overlap=max(lista_overlaps_sin_nueves)
                    
                    # promedio de gaps y mas
                    #lista_gaps_sin_nueves=[]
                    #for i in range(0,len(fil_time["fecha"])):
                    #    if fil_time["num_gaps"].iloc[i] != -9: #se le puso -9 a los datos no disponibles
                    #        lista_gaps_sin_nueves.append(fil_time["num_gaps"].iloc[i])
                    #gaps_prom = round(np.mean(np.array(lista_gaps_sin_nueves)),2)
                    #num_gaps = np.sum(lista_gaps_sin_nueves)
                    #max_gaps = max(lista_gaps_sin_nueves)
                    
                    #limites eje auxiliar o secundario
                    #fig2.update_yaxes(range=(0,int(max_overlap)*1.5),secondary_y=True) #Limites en eje conteo
                    #fig2.update_yaxes(range=(0,int(max_gaps)*2),title_text="Conteo (Overlaps, Gaps)",secondary_y=True)  #Limites en eje conteo

                    #detalles
                    tit2 = 'Offset ' + estacion + " "+sensor+can+" | "+fecha_ini.strftime("%Y-%m-%d")+" - "+fecha_fin.strftime("%Y-%m-%d")
                    fig2.update_layout(width = 800, height = 310,title_text=tit2,
                                    legend=dict(orientation="h",yanchor="bottom",y=1.1,xanchor="right",x=1),
                                    yaxis_title='Offset')
                    
                    fig2.update_xaxes(showspikes=True,spikethickness=1,spikesnap="cursor",spikemode="across")#este es para colocar las gias al pasar el cursor
                    fig2.update_yaxes(showspikes=True,spikethickness=1,spikesnap="cursor",spikemode="across")#este es para colocar las gias al pasar el cursor
                    
                    fig2.update_xaxes(showgrid=True, ticklabelmode="period",tickformat="%d\n%b\n%Y", gridwidth=1, gridcolor='LightGrey')
                    fig2.update_xaxes(showline=True, linewidth=2, linecolor='lightgray', mirror=True)
                    fig2.update_yaxes(showline=True, linewidth=2, linecolor='lightgray', mirror=True, )
                    
                    st.plotly_chart(fig2, use_container_width=True)

                with col2:
                            
                    #st.write(f"Max. offset : {max_offset}" )
                    
                    st.write(".")
                    st.markdown(f"**:red[Min. offset:]** {min_offset} **:red[| Max. offset :]** {max_offset}" )
                    st.markdown(f"**:red[Promedio offset:]** {offset_prom}" )
                    
                    #st.write(f"Max. gaps: {max_gaps}")
                    #st.markdown(f"**Num. gaps:** {num_gaps} **| Max. gaps:** {max_gaps}")
                    #st.markdown(f"**Promedio de gaps:** {gaps_prom}")
                    
                    #st.write(f"Max. overlaps: {max_overlap}")
                    #st.markdown(f"**:green[Num. overlaps:]** {num_overlaps} **:green[| Max. overlaps:]** {max_overlap}")
                    #st.markdown(f"**:green[Promedio overlaps:]** {overlaps_prom}")
                
                return  [min_offset,max_offset,offset_prom, codigo_l]
            
            else:
                st.error("No hay datos de esta estación en estas fechas")
                return  [0,0,0, codigo_l]

    def plot_gapover_din(self, can):

        #folder = self.folder
        estacion = self.estacion
        sensor = self.sensor
        fecha_ini = self.fecha_ini
        fecha_fin = self.fecha_fin
        list_est = self.list_est
        fil_time = self.fil_time
        dias_ad = self.dias_ad 
        
        codigo_l = list_est[0][-10:-8]
        if len(fil_time) != 0:
            
            
            #-----GRAFICA CONTEO DE GAPS, OVERLAPS y offset
            #-----GRAFICA OFFSET
            
            col1,col2= st.columns([5,1.1])
            with col1:
                ###### Esto es para que el valor de offset despues de l 0 sea 0 tambien
                busca_ceros=[]
                i=0
                
                #for e in fil_time["offs"]:
                    
                #    if e == 0:
                #        busca_ceros.append(i)
                #    if len(busca_ceros)>0:
                #        if e!=0:
                            
                #            fil_time["offs"].iloc[i]=0
                #            busca_ceros=[]
                #    if e == -9:
                #        fil_time["offs"].iloc[i]=None
                    
                #    i=i+1
                
                ###Datos promedio, y mas
                #lista_offset_sin_ceros=[]
                #for of in fil_time["offs"]:
                #    if of != 0 and str(of) != "nan":
                        
                #        lista_offset_sin_ceros.append(of)
                        
                
                
                #if len(lista_offset_sin_ceros) >0:
                #    offset_prom=round(np.mean(np.array(lista_offset_sin_ceros)),2) #Promedio del offset
                #    max_offset=round(max(lista_offset_sin_ceros),2)
                #    min_offset=round(min(lista_offset_sin_ceros),2)
                #else:
                #    offset_prom=0   
                #    max_offset=""
                #    min_offset=""
                ######
                
                #------para mejorar visalizacion offset. Los ceros puestos que son cuando no se pudo calcular (por no disponibilidad)
                #los convierte en el promedio. Si existen valores gigantes, 50 veces mayor a la media grafica el offset en escala log
                #escala="linear"
                #i=0       
                #for e in fil_time["offs"]:
                #    if e==0:
                #        fil_time["offs"].iloc[i]=offset_prom
                #    if abs(e) >abs(50*offset_prom) or abs(e)>100000:
                        #escala="log"
                #        escala="linear"
                #    i=i+1
                #-------- 
                
                fig4 = make_subplots(specs=[[{"secondary_y": True}]])
                
                
                #if escala == "linear":
                #    fig2.add_trace(
                #        go.Scatter( x=fil_time["fecha"], y=fil_time["offs"], mode = "lines",marker_color='rgba(225,0,0,1)', connectgaps=False,name='offset',
                #                   hovertemplate="Fecha: %{x}<br>" +"Offset: %{y}<br>" +"<extra></extra>",  #hovertemplate, para seleccionar el texto al mostrar al pasar el cursor
                #                   hoverlabel=dict(bgcolor="#C11515",font_size=15,font_color="white",font_family="Times New Roman")),secondary_y=False,) #hoverlabel para configurar la forma del texto de información
                #    fig2.add_hline(y=offset_prom,line_dash="dot",line_width=1,line_color="grey",annotation_text=f"media: {offset_prom}", 
                #    annotation_position="bottom")
                    #fig2.update_yaxes(type="log")

                #if escala == "log": #si esta en escala lineal obtiene el abs del offset (para poder graficar valores negativos del offset original, no                     grafica media porque pierde sentido, esto sucede porque hay valores gigantes respecto a la media)
                #    fig2.add_trace(go.Scatter( x=fil_time["fecha"], y=np.abs(np.array(list(fil_time["offs"]))), mode = "lines",marker_color='rgba(225,0,0,1)', connectgaps=True,name='offset_abs',),secondary_y=False,)
                #    fig2.add_trace(go.Scatter( x=[f_i], y=[0], mode = "lines",marker_color='rgba(140,140,140,1)',name="Gaps",line=dict(width=9),), secondary_y=False)
                
                #leyenda overlaps
                fig4.add_trace(
                    go.Scatter( x=[fecha_ini ], y=[0], mode = "lines",marker_color='rgba(50,205,50,0.7)',name="Overlaps",line=dict(width=2),),secondary_y=False,) #hoverlabel para configurar la forma del texto de información
                #leyenda de gaps
                fig4.add_trace(
                    go.Scatter( x=[fecha_ini ], y=[0], mode = "lines",marker_color='rgba(71,71,71,0.7)',name="Gaps",line=dict(width=2),),secondary_y=False,) #hoverlabel para configurar la forma del texto de información

                #para graficar conteo de overlaps
                for n_over,n_gaps,time in zip(fil_time["num_overlaps"],fil_time["num_gaps"],fil_time["fecha"]):
                    
                    #graficar numero de overlaps por dia
                    fig4.add_trace(go.Scatter(x=[time,time], y=[0,n_over],text=[0,n_gaps],mode='lines', connectgaps=False,marker_color='rgba(50,205,50,0.7)',name='Overlaps',showlegend=False   ,line=dict(width=2),
                                hovertemplate="Fecha: %{x}<br>" +"Overlaps: %{y}<br>"+"Gaps: %{text}<br>" +"<extra></extra>",  #hovertemplate, para seleccionar el texto al mostrar al pasar el cursor
                                hoverlabel=dict(bgcolor="#34C115",font_size=15,font_color="white",font_family="Times New Roman")),secondary_y=False,) #hoverlabel para configurar la forma del texto de información
                    #graficar numero de gaps por dia
                    fig4.add_trace(go.Scatter(x=[time,time], y=[0,n_gaps],text=[0,n_over],mode='lines', connectgaps=False,marker_color='rgba(71,71,71,0.7)',name='Gaps',showlegend=False   ,line=dict(width=2),
                                hovertemplate="Fecha: %{x}<br>" +"Gaps: %{y}<br>"+"Overlaps: %{text}<br>" +"<extra></extra>",  #hovertemplate, para seleccionar el texto al mostrar al pasar el cursor
                            hoverlabel=dict(bgcolor="#797A79",font_size=15,font_color="white",font_family="Times New Roman")),secondary_y=False,) #hoverlabel para configurar la forma del texto de información
                
                ##Datos promedio y mas
                lista_overlaps_sin_nueves=[]
                for i in range(0,len(fil_time["fecha"])):
                    if np.isnan(fil_time["num_overlaps"].iloc[i]):
                        continue
                    if fil_time["num_overlaps"].iloc[i] != -9: #se le puso -9 a los datos no disponibles
                        lista_overlaps_sin_nueves.append(fil_time["num_overlaps"].iloc[i])
                        
                if len(lista_overlaps_sin_nueves)!=0:
                    overlaps_prom=round(np.mean(np.array(lista_overlaps_sin_nueves)),2)   #Promedio overlaps
                    num_overlaps = np.sum(lista_overlaps_sin_nueves)
                    max_overlap=max(lista_overlaps_sin_nueves)
                if len(lista_overlaps_sin_nueves)==0:
                    overlaps_prom=0
                    num_overlaps=0
                    max_overlap=0
                # promedio de gaps y mas
                lista_gaps_sin_nueves=[]
                for i in range(0,len(fil_time["fecha"])):
                    if np.isnan(fil_time["num_gaps"].iloc[i]):
                        continue
                    if fil_time["num_gaps"].iloc[i] != -9: #se le puso -9 a los datos no disponibles
                        lista_gaps_sin_nueves.append(fil_time["num_gaps"].iloc[i])
                
                if len(lista_gaps_sin_nueves) !=0:
                    gaps_prom = round(np.mean(np.array(lista_gaps_sin_nueves)),2)
                    num_gaps = np.sum(lista_gaps_sin_nueves)
                    max_gaps = max(lista_gaps_sin_nueves)
                if len(lista_gaps_sin_nueves) ==0:
                    gaps_prom=0
                    num_gaps=0
                    max_gaps=0
                #limites eje auxiliar o secundario
                fig4.update_yaxes(range=(0,int(max_overlap)),secondary_y=False) #Limites en eje conteo
                fig4.update_yaxes(range=(0,int(max_gaps)),title_text="Conteo (Overlaps, Gaps)",secondary_y=False)  #Limites en eje conteo

                #detalles
                tit2 = 'Conteo de Gaps y Overlaps' + estacion + " "+sensor+can+" | "+fecha_ini.strftime("%Y-%m-%d")+" - "+fecha_fin.strftime("%Y-%m-%d")
                fig4.update_layout(width = 800, height = 310,title_text=tit2,)
                
                fig4.update_xaxes(showspikes=True,spikethickness=1,spikesnap="cursor",spikemode="across")#este es para colocar las gias al pasar el cursor
                fig4.update_yaxes(showspikes=True,spikethickness=1,spikesnap="cursor",spikemode="across")#este es para colocar las gias al pasar el cursor
                
                fig4.update_xaxes(showgrid=True, ticklabelmode="period",tickformat="%d\n%b\n%Y", gridwidth=1, gridcolor='LightGrey')
                fig4.update_xaxes(showline=True, linewidth=2, linecolor='lightgray', mirror=True)
                fig4.update_yaxes(showline=True, linewidth=2, linecolor='lightgray', mirror=True, )
                
                st.plotly_chart(fig4, use_container_width=True)

            with col2:
                        
                #st.write(f"Max. offset : {max_offset}" )
                
                st.write(".")
                #st.markdown(f"**:red[Min. offset:]** {min_offset} **:red[| Max. offset :]** {max_offset}" )
                #st.markdown(f"**:red[Promedio offset:]** {offset_prom}" )
                
                #st.write(f"Max. gaps: {max_gaps}")
                st.markdown(f"**Num. gaps:** {num_gaps} **| Max. gaps:** {max_gaps}")
                st.markdown(f"**Promedio de gaps:** {gaps_prom}")
                
                #st.write(f"Max. overlaps: {max_overlap}")
                st.markdown(f"**:green[Num. overlaps:]** {num_overlaps} **:green[| Max. overlaps:]** {max_overlap}")
                st.markdown(f"**:green[Promedio overlaps:]** {overlaps_prom}")            

            return [num_gaps,max_gaps,gaps_prom,num_overlaps,max_overlap,overlaps_prom, codigo_l]
        
        else:
            st.error("No hay datos de esta estación en estas fechas")
            return [0,0,0,0,0,0, codigo_l]
        
    def plot_ppsd_din(self, can):
        
        #folder = self.folder
        estacion = self.estacion
        sensor = self.sensor
        fecha_ini = self.fecha_ini
        fecha_fin = self.fecha_fin
        list_est = self.list_est
        fil_time = self.fil_time
        dias_ad = self.dias_ad 
        codigo_l = list_est[0][-10:-8]
        
        if len(fil_time) != 0:
            
            
            
            #-------------GRAFICA DE PPSD, Y CONTEO PICOS E HORAS
            
            col1,col2= st.columns([5,1.1])
            with col1:
                lista_ppsd_sin_menosnueves=[]
                lista_ppsd=[]
                for of in fil_time["p_ppsd"]:
                    if np.isnan(of):
                        continue
                    if of != -9:
                        lista_ppsd_sin_menosnueves.append(of)
                        lista_ppsd.append(of)
                        
                    if of == -9:
                        lista_ppsd.append(None)
                        
                fig3 = make_subplots(specs=[[{"secondary_y": True}]])
                fig3.add_trace(
                    go.Scatter( x=fil_time["fecha"], y=lista_ppsd, mode = "lines",marker_color='rgba(13,144,0,1)', connectgaps=False,name='ppsd',
                            hovertemplate="Fecha: %{x}<br>" +"ppsd: %{y}%<br>" +"<extra></extra>",  #hovertemplate, para seleccionar el texto al mostrar al pasar el cursor
                            hoverlabel=dict(bgcolor="#34C115",font_size=15,font_color="white",font_family="Times New Roman")),secondary_y=False,)
                
                #leyenda de picos
                fig3.add_trace(go.Scatter( x=[fecha_ini ], y=[0], mode = "lines",marker_color='rgba(128,0,255,0.7)',name="Picos",line=dict(width=2),), secondary_y=True)

                #para graficar conteo de picos
                for n_picos,time in zip(fil_time["peaks"],fil_time["fecha"]):

                    fig3.add_trace(
                        go.Scatter(x=[time,time], y=[0,n_picos],mode='lines', connectgaps=False,marker_color='rgba(128,0,255,0.7)',name='Picos',showlegend=False   ,line=dict(width=2),
                            hovertemplate="Fecha: %{x}<br>" +"Picos: %{y}<br>" +"<extra></extra>",  #hovertemplate, para seleccionar el texto al mostrar al pasar el cursor
                            hoverlabel=dict(bgcolor="#8B41B0",font_size=15,font_color="white",font_family="Times New Roman")),  secondary_y=True,)

                #lista sin nueves
                lista_picos_sin_nueves=[]
                for i in range(0,len(fil_time["fecha"])):
                    if np.isnan(fil_time["peaks"].iloc[i]):
                        continue
                    if fil_time["peaks"].iloc[i] != -9: #se le puso -9 a los datos no disponibles
                        lista_picos_sin_nueves.append(fil_time["peaks"].iloc[i])     
                #promedio y mas 
                if len(lista_ppsd_sin_menosnueves) != 0:
                    ppsd_prom = round(np.mean(np.array(lista_ppsd_sin_menosnueves)),2)
                if len(lista_ppsd_sin_menosnueves) == 0:
                    ppsd_prom=None
                if len(lista_picos_sin_nueves) != 0:
                    num_picos = np.sum(lista_picos_sin_nueves)
                    max_picos = max(lista_picos_sin_nueves)
                if len(lista_picos_sin_nueves) == 0:
                    num_picos=0
                    max_picos=0
                #limites eje auxiliar o secundario
                fig3.update_yaxes(range=(0,int(max_picos)),title_text="Conteo Picos",secondary_y=True)  #Limites en eje conteo

                #detalles
                tit3 = '% ppsd fuera - picos ' + estacion + " "+sensor+can+" | "+fecha_ini.strftime("%Y-%m-%d")+" - "+fecha_fin.strftime("%Y-%m-%d")
                fig3.update_layout(width = 800, height = 310,title_text=tit3, 
                                legend=dict(orientation="h",yanchor="bottom",y=1.1,xanchor="right",x=1),
                                yaxis_title='ppsd (%)')
                
                fig3.update_xaxes(showspikes=True,spikethickness=1,spikesnap="cursor",spikemode="across")#este es para colocar las gias al pasar el cursor
                fig3.update_yaxes(showspikes=True,spikethickness=1,spikesnap="cursor",spikemode="across")#este es para colocar las gias al pasar el cursor
                
                fig3.update_xaxes(showgrid=True, ticklabelmode="period",tickformat="%d\n%b\n%Y", gridwidth=1, gridcolor='LightGrey')
                fig3.update_xaxes(showline=True, linewidth=2, linecolor='lightgray', mirror=True)
                fig3.update_yaxes(showline=True, linewidth=2, linecolor='lightgray', mirror=True, )
                
                st.plotly_chart(fig3, use_container_width=True)
            
            with col2:
                
                st.write(".") 
                st.write("_________________________")    
                st.markdown(f"**:violet[Num. picos:]** {num_picos} **:violet[| Max. de picos:]** {max_picos}")
                st.markdown(f"**:green[Promedio (%) ppsd por fuera:]** {ppsd_prom}%")
                st.write("_________________________")
                
            return [ppsd_prom,num_picos,max_picos, codigo_l]
        
        else:
            st.error("No hay datos de esta estación en estas fechas")  
            return [0,0,0, codigo_l] 
             
    def plt_ppsd_dia(self, can):
        
        fecha_ini = self.fecha_ini
        fecha_fin = self.fecha_fin
        
        with open("ARGC.json","r") as json_file: 
            data_ppsd_r = json.load(json_file)
        #print("#len ppsd",len(data_ppsd_r[0][0]))
        with st.container():

            f = st.slider('Fecha PPSD', fecha_ini, fecha_fin)
            
            print(data_ppsd_r[0]["fecha"])
            for e in range(len(data_ppsd_r[0]["fecha"])):
                fecha = data_ppsd_r[0]["fecha"][e]
                
                a = st.write(fecha)
        
    

def plot_ruido(fecha,est,sensor, can,red="CM"):#yyyy-mm-dd
    
    folder_ruido = os.path.dirname(os.path.abspath(__file__))+f'/Datos/{red}/ppsd_mean/'
    
    
    if os.path.exists(folder_ruido+f"{est}_{sensor}{can}.json") == True:
        with open(folder_ruido+f"{est}_{sensor}{can}.json", 'r') as file:
            ruido_json = json.load(file)
        
        for e in range(len(ruido_json[0]["fecha"])):
            
            if ruido_json[0]["fecha"][e] == fecha:
                fecha_ruido= ruido_json[0]["fecha"][e]
    
                fig5 = make_subplots(specs=[[{"secondary_y": False}]])
                fig5.add_trace(
                        go.Scatter( x=ruido_json[0]["frecuencias"][e], y=ruido_json[0]["mean"][e][::-1], mode = "lines",marker_color='rgba(13,144,0,1)', connectgaps=False,name='ppsd',),secondary_y=False)
                
                HNM_valores=[-129.3,-130.5,-131.5,-134.5,-137.5,-125.5,-118,-115.5,-97.5,-100,-107,-112,-115,-115,-111,-102,-96.5,-95,-93,-91.5]
                LNM_valores= [-185.2,-185,-185,-187.5,-180.1,-165.1,-165.6,-163.6,-141.1,-144.3,-152.8,-160.5,-165.1,-167.5,-166.7,-166.7,-166.7,-166.7,-167.4,-168]
                petter_frec=[0.006,0.008,0.01,0.02,0.04,0.06,0.08,0.1,0.2,0.3,0.5,0.7,0.9,2,3,4,5,6,8,10]

                fig5.add_trace(
                    go.Scatter( x=petter_frec, y=HNM_valores, mode = "lines",marker_color='rgba(59,59,59,1)',name="HNM",line=dict(width=3),),secondary_y=False,) #hoverlabel para configurar la forma del texto de información
                    
                fig5.add_trace(
                    go.Scatter( x=petter_frec, y=LNM_valores, mode = "lines",marker_color='rgba(59,59,59,1)',name="LNM",line=dict(width=3),), secondary_y=False)

                fig5.update_layout(xaxis_range=[np.log10(0.006), np.log10(100)], yaxis_range=[-200, -50],width=300,height=400)
                fig5.update_xaxes(type="log")
                
                fig5.update_xaxes(showgrid=True, ticklabelmode="period", gridwidth=1, gridcolor='LightGrey')
                fig5.update_yaxes(showgrid=True, ticklabelmode="period", gridwidth=1, gridcolor='LightGrey')
                
                fig5.update_xaxes(showline=True, linewidth=2, linecolor='gray', mirror=True,title_text="Frecuencias (Hz)")
                fig5.update_yaxes(showline=True, linewidth=2, linecolor='gray', mirror=True,title_text="dB" )
                
                
                st.plotly_chart(fig5, use_container_width=True)
                
def actualizar_est_siigeo():
    file_xlsx = os.path.dirname(os.path.abspath(__file__))+"/dat_est/stationReportCodLocation.xlsx"
    file = requests.get(f"https://siigeo.sgc.gov.co/stations/stationExcelCodLocation/", verify=False) 
    open(file_xlsx, 'wb').write(file.content)

    file_uv_xlsx = os.path.dirname(os.path.abspath(__file__))+"/dat_est/lastServiceMaintenance.xlsx"

    file = requests.get("https://siigeo.sgc.gov.co/reports/lastServiceMaintenance", verify=False) 
    open(file_uv_xlsx, 'wb').write(file.content)
    
def get_data_est(estacion, actualizar):
        
        file_xlsx = os.path.dirname(os.path.abspath(__file__))+"/dat_est/stationReportCodLocation.xlsx"
        
        tab_est = pd.read_excel(file_xlsx)
        
        #filtro por Estación
        filtro_estacion = (tab_est["IDENTIFICADOR"] == estacion)
        fil_est = tab_est[filtro_estacion]
        
        if len(fil_est) > 0:
            
            
            nombre = fil_est["NOMBRE"].iloc[0]
            departamento = fil_est["DEPARTAMENTO"].iloc[0]
            municipio = fil_est["MUNICIPIO"].iloc[0]
            latitud = fil_est["LATITUD (°)"].iloc[0]
            longitud = fil_est["LONGITUD (°)"].iloc[0]
            altura = fil_est["ELEVACION (msnm)"].iloc[0]
            
            resp_elect =fil_est["ELECTRÓNICO RESPONSABLE"].iloc[0]
            resp_tem = fil_est["TEMÁTICO RESPONSABLE"].iloc[0]
            
            estado = fil_est["ESTADO ESTACIÓN"].iloc[0]
            cond_ins = fil_est["CONDICIÓN DE INSTALACIÓN"].iloc[0]
            transmi = fil_est["TIPO DE TRANSMISIÓN"].iloc[0]
            adquisicion = fil_est["TIPO DE ADQUISICIÓN"].iloc[0]
            tip_est = fil_est["TIPO DE ESTACIÓN"].iloc[0]
            
            
            
            n_sensores = len(fil_est)
            col1, col2 = st.columns([1,1])
            with col1:
                st.info(f"""
                            **Nombre :** {nombre}\n
                            **Departamento :** {departamento} | **Municipio :** {municipio}\n
                            **Latitud :** {latitud}° | **Longitud :** {longitud}° | **Altura :** {altura} m\n
                            **Electronico responsable :** {resp_elect}\n
                            **Sismologo responsable :** {resp_tem}""")
                #st.write("____")
                st.info(f"""
                        **Estado :** {estado}\n
                        **Tipo de transmisión :** {transmi} | **Tipo de adquisición :** {adquisicion}\n
                        **Condición de instalación :** {cond_ins} | **Tipo de estación:** {tip_est}\n
                        """)
            with col2:
                for e in range(n_sensores):
                    if np.isnan(fil_est["CÓDIGO LOCALIZACION"].iloc[e]) == False: 
                        cod_loc = int(fil_est["CÓDIGO LOCALIZACION"].iloc[e])
                    else:
                        cod_loc = "--"
                    equipos = fil_est["INSTRUMENTACIÓN"].iloc[e]
                    f_in = fil_est["FECHA INICIO COD."].iloc[e]
                    f_fi = fil_est["FECHA FIN COD."].iloc[e]
                    #st.markdown(f"**Sensor {cod_loc}**")
                    st.success(f"""
                               **Sensor {str(cod_loc).rjust(2,"0")}**\n
                            **Sensor y digitalizador :** {equipos}\n
                            **Fecha inicio :** {f_in} | **Fecha fin :** {f_fi}""")
                    #st.write("____")

                ##ultima visita
                file_xlsx = os.path.dirname(os.path.abspath(__file__))+"/dat_est/lastServiceMaintenance.xlsx"
                tab_est = pd.read_excel(file_xlsx)
                tab_est2 = tab_est.drop(tab_est.index[[0]])
                tit=["ESTACION","ESTADO ESTACION","FECHA MANTENIMIENTO","TIPO DE MANTENIMIENTO", "RED DE MONITOREO","COMISIONADOS",	"REGISTRO CARGADO POR",	"COMENTARIOS"]
                tab_est2.columns = tit

                #filtro por Estación
                filtro_estacion = (tab_est2["ESTACION"] == estacion)
                fil_est3 = tab_est2[filtro_estacion]
                #st.dataframe(fil_est3)
                
                fecha_man = fil_est3["FECHA MANTENIMIENTO"].iloc[0]
                tip_man=fil_est3["TIPO DE MANTENIMIENTO"].iloc[0].title()
                vis_por = fil_est3["COMISIONADOS"].iloc[0].title()
                comentarios = fil_est3["COMENTARIOS"].iloc[0].lower().replace("\n"," ")

                st.warning(f"""
                **Última visita**\n
                **Fecha:** {fecha_man} | **Tipo:** {tip_man}\n
                **Comisionados:** {vis_por}\n
                **Comentarios:** {comentarios}""")

                    
            
        else:
            st.error("**No hay información de la estación en SIIGEO**",icon="🚨")
        #st.dataframe(fil_est)


def actualizar_bitacora_siigeo():
    file_xlsx = os.path.dirname(os.path.abspath(__file__))+"/dat_est/reportHistoricalBitacora.xlsx"
    year1="2022"
    month1="03"
    day1="08"

    time_now=datetime.now()
    year2=time_now.year
    month2=time_now.month
    
    #day2=time_now.day+1
    d = time_now+relativedelta(days=1)
    day2=d.day
    file = requests.get(f"https://siigeo.sgc.gov.co/virtualLogbook/reportExcelNovelty/{year1}-{month1}-{day1},{year2}-{month2}-{day2}", verify=False)
    print("####AQUII, SIIGEO",f"https://siigeo.sgc.gov.co/virtualLogbook/reportExcelNovelty/{year1}-{month1}-{day1},{year2}-{month2}-{day2}" ) 
    open(file_xlsx, 'wb').write(file.content)

def get_bitacora(estacion, fecha_ini, fecha_fin):


    file_xlsx = os.path.dirname(os.path.abspath(__file__))+"/dat_est/reportHistoricalBitacora.xlsx"
        
    tab_est = pd.read_excel(file_xlsx)
    tab_est2 = tab_est.drop(tab_est.index[[0,1,2,3,4]])
    tit=["FECHA","TURNO","CATEGORIA","ESTACION","NOVEDAD","OBSERVACIONES"]
    tab_est2.columns = tit


    #arreglar fecha de la tabla
    f_date = []
    for e in tab_est2["FECHA"]:
        year, month, day = e.split("-")
        f_date.append(datetime(int(year),int(month),int(day)))
    
    #Cambio de formato de fecha times_gaps a datetime
    tab_est2["FECHA"] = f_date
    #filtro por Estación y fechas
    filtro_estacion = (tab_est2["ESTACION"] == estacion)  & (tab_est2["FECHA"] >= fecha_ini) & (tab_est2["FECHA"] <= fecha_fin) 
    fil_est = tab_est2[filtro_estacion]

    if len(fil_est) > 0:

        #for e in range(len(fil_est)):
        #    fech = fil_est["FECHA"].iloc[e]
        #    turno = fil_est["TURNO"].iloc[e]
        #    novedad = str(fil_est["NOVEDAD"].iloc[e])
        #    obs=str(fil_est["OBSERVACIONES"].iloc[e]).replace("{", "[").replace("}", "]")


        #    st.info(f"""
        #                    **Fecha :** {fech} | **Turno :** {turno}\n
        #                    **Novedad :** {novedad}\n 
        #                    **Observaciones :** {obs}""")

        #st.dataframe(fil_est.loc[:,["FECHA","TURNO","NOVEDAD","OBSERVACIONES"]])
        st.dataframe(fil_est.loc[:,["FECHA","NOVEDAD","OBSERVACIONES"]], column_config={"NOVEDAD":st.column_config.Column(width="large",)},hide_index=True)

def get_ultima_v(estacion,sensor, folder,option_cod_sen):

    file_xlsx = os.path.dirname(os.path.abspath(__file__))+"/dat_est/lastServiceMaintenance.xlsx"
        
    tab_est = pd.read_excel(file_xlsx)
    tab_est2 = tab_est.drop(tab_est.index[[0]])
    tit=["ESTACION","ESTADO ESTACION","FECHA MANTENIMIENTO","TIPO DE MANTENIMIENTO", "RED DE MONITOREO","COMISIONADOS",	"REGISTRO CARGADO POR",	"COMENTARIOS"]
    tab_est2.columns = tit

    
    #filtro por Estación
    filtro_estacion = (tab_est2["ESTACION"] == estacion)
    fil_est3 = tab_est2[filtro_estacion]
    #st.dataframe(fil_est3)
    
    fecha_man = fil_est3["FECHA MANTENIMIENTO"].iloc[0]
    tip_man=fil_est3["TIPO DE MANTENIMIENTO"].iloc[0].title()
    vis_por = fil_est3["COMISIONADOS"].iloc[0].title()
    comentarios = fil_est3["COMENTARIOS"].iloc[0].lower()


    #folder_semestre = os.path.dirname(os.path.abspath(__file__))+f"/Inf_semestrales/{estacion}_{sensor}/{semestre}/"
    folder_semestre = folder
    if option_cod_sen == ".": option_cod_sen=""
    if len(fil_est3)>0:
        with open(folder_semestre + f"inf_{estacion}_{sensor}{option_cod_sen}.json", 'r') as file:
            inf_json_ul = json.load(file)
        
        #fecha_str= f"{fecha_man.year}-{fecha_man.month}-{fecha_man.day}"
        inf_json_ul[0]["ultima_v"][0] = fecha_man
        inf_json_ul[0]["ultima_v"][1] =tip_man
        if len(comentarios) != 0:
            inf_json_ul[0]["ultima_v"][2]=comentarios
        if len(comentarios) == 0:
            inf_json_ul[0]["ultima_v"][2]=inf_json_ul[0]["ultima_v"][2]

        if len(vis_por) != 0:
            inf_json_ul[0]["ultima_v"][3]=vis_por
        if len(vis_por) == 0:
            inf_json_ul[0]["ultima_v"][3]=inf_json_ul[0]["ultima_v"][3]
        
        with open(folder_semestre + f"inf_{estacion}_{sensor}{option_cod_sen}.json", 'w') as file:
                json.dump(inf_json_ul, file, indent=4)

    

def actualizar_ultima_v():

    file_xlsx = os.path.dirname(os.path.abspath(__file__))+"/dat_est/lastServiceMaintenance.xlsx"

    file = requests.get("https://siigeo.sgc.gov.co/reports/lastServiceMaintenance", verify=False) 
    open(file_xlsx, 'wb').write(file.content)


def create_info_s(est,sensor,fecha_ini, fecha_fin,folder,semestre):

    folder_semestre = os.path.dirname(os.path.abspath(__file__))+f"/Inf_semestrales/{est}_{sensor}/{semestre}/"
    if os.path.exists(folder_semestre) == False:
        os.makedirs(folder_semestre)

        for com in ["Z","N","E"]:

            try:
                
                
                list_est_sen = glob.glob(os.path.dirname(os.path.abspath(__file__))+f"/Datos/CM/Tablas_Cal_tiempos/{est}_{sensor}Z*")
                cod_sensor=[]
                if len(list_est_sen) == 2:
                    for s in list_est_sen:
                        if s[-10:-8] not in cod_sensor:
                            cod_sensor.append(s[-10:-8])
                    for cod_s in cod_sensor:
                        file = glob.glob(os.path.dirname(os.path.abspath(__file__))+f"/Datos/CM/Tablas_Cal_tiempos/{est}_{sensor}{com}_{cod_s}*")
                        
                        plt_disp = Plot(est,sensor, fecha_ini,fecha_fin,file)
                        plt_disp.plot_disp_fij(com,folder_semestre,two_sensores=True)
                        plt_disp.plot_gapover_fij(com, folder_semestre, two_sensores=True)
                        plt_disp.plot_off_fij(com, folder_semestre,two_sensores=True)
                        plt_disp.plot_ppsd_fij(com, folder_semestre,two_sensores=True)
                
                if len(list_est_sen) == 1:
                    
                    file = glob.glob(os.path.dirname(os.path.abspath(__file__))+f"/Datos/CM/Tablas_Cal_tiempos/{est}_{sensor}{com}*")
                    
                    plt_disp = Plot(est,sensor, fecha_ini,fecha_fin,file)
                    plt_disp.plot_disp_fij(com,folder_semestre)
                    plt_disp.plot_gapover_fij(com, folder_semestre)
                    plt_disp.plot_off_fij(com, folder_semestre)
                    plt_disp.plot_ppsd_fij(com, folder_semestre)
                    
                
                #plt_disp.plot_disp_din(com,save=True, folder=folder_semestre) #Disponibilidad
                
            
            except:
                print("No hay componentes horizontales")
                continue
            
        hoy = datetime.now()
        fecha_hoy= f"{hoy.year}-{hoy.month}-{hoy.day}"
        #Creacion de Json
        datos_json = [{'funcionamiento':" ",'disponibilidad':" ",'gapover':" ", 'calidad':" ", 'offset':" ","ruido":" ","ultima_v":["","","",""],"recomendaciones":" ","fecha_creacion":f"{fecha_hoy} ","n_img_ruido":0,"n_img_recom":0}]
        
        list_est_sen = glob.glob(os.path.dirname(os.path.abspath(__file__))+f"/Datos/CM/Tablas_Cal_tiempos/{est}_{sensor}Z*")
        cod_sensor=[]
        
        if len(list_est_sen) == 2:
            for s in list_est_sen:
                if s[-10:-8] not in cod_sensor:
                    cod_sensor.append(s[-10:-8])
            for cod_s in cod_sensor:
                         
                with open(folder_semestre + f"inf_{est}_{sensor}{cod_s}.json", 'w') as (file):
                    json.dump(datos_json, file, indent=4)
        if len(list_est_sen) == 1:
            with open(folder_semestre + f"inf_{est}_{sensor}.json", 'w') as (file):
                json.dump(datos_json, file, indent=4)

def create_info_s_all(est,sensor,fecha_ini, fecha_fin,folder,semestre):
    #
    folder_semestre_all = os.path.dirname(os.path.abspath(__file__))+f"/Inf_semestrales/{est}/{semestre}/"

    #
    folder_semestre = os.path.dirname(os.path.abspath(__file__))+f"/Inf_semestrales/{est}_{sensor}/{semestre}/"
    if os.path.exists(folder_semestre) == False:
        os.makedirs(folder_semestre)

        for com in ["Z","N","E"]:

            try:
                
                
                list_est_sen = glob.glob(os.path.dirname(os.path.abspath(__file__))+f"/Datos/CM/Tablas_Cal_tiempos/{est}_{sensor}Z*")
                cod_sensor=[]
                if len(list_est_sen) == 2:
                    for s in list_est_sen:
                        if s[-10:-8] not in cod_sensor:
                            cod_sensor.append(s[-10:-8])
                    for cod_s in cod_sensor:
                        file = glob.glob(os.path.dirname(os.path.abspath(__file__))+f"/Datos/CM/Tablas_Cal_tiempos/{est}_{sensor}{com}_{cod_s}*")
                        
                        plt_disp = Plot(est,sensor, fecha_ini,fecha_fin,file)
                        plt_disp.plot_disp_fij(com,folder_semestre,two_sensores=True)
                        plt_disp.plot_gapover_fij(com, folder_semestre, two_sensores=True)
                        plt_disp.plot_off_fij(com, folder_semestre,two_sensores=True)
                        plt_disp.plot_ppsd_fij(com, folder_semestre,two_sensores=True)
                
                if len(list_est_sen) == 1:
                    
                    file = glob.glob(os.path.dirname(os.path.abspath(__file__))+f"/Datos/CM/Tablas_Cal_tiempos/{est}_{sensor}{com}*")
                    
                    plt_disp = Plot(est,sensor, fecha_ini,fecha_fin,file)
                    plt_disp.plot_disp_fij(com,folder_semestre)
                    plt_disp.plot_gapover_fij(com, folder_semestre)
                    plt_disp.plot_off_fij(com, folder_semestre)
                    plt_disp.plot_ppsd_fij(com, folder_semestre)
                    
                
                #plt_disp.plot_disp_din(com,save=True, folder=folder_semestre) #Disponibilidad
                
            
            except:
                print("No hay componentes horizontales")
                continue
            
        hoy = datetime.now()
        fecha_hoy= f"{hoy.year}-{hoy.month}-{hoy.day}"
        #Creacion de Json
        datos_json = [{'funcionamiento':" ",'disponibilidad':" ",'gapover':" ", 'calidad':" ", 'offset':" ","ruido":" ","ultima_v":["","","",""],"recomendaciones":" ","fecha_creacion":f"{fecha_hoy} ","n_img_ruido":0,"n_img_recom":0}]
        
        list_est_sen = glob.glob(os.path.dirname(os.path.abspath(__file__))+f"/Datos/CM/Tablas_Cal_tiempos/{est}_{sensor}Z*")
        cod_sensor=[]
        
        if len(list_est_sen) == 2:
            for s in list_est_sen:
                if s[-10:-8] not in cod_sensor:
                    cod_sensor.append(s[-10:-8])
            for cod_s in cod_sensor:
                         
                with open(folder_semestre + f"inf_{est}_{sensor}{cod_s}.json", 'w') as (file):
                    json.dump(datos_json, file, indent=4)
        if len(list_est_sen) == 1:
            with open(folder_semestre + f"inf_{est}_{sensor}.json", 'w') as (file):
                json.dump(datos_json, file, indent=4)




def create_info_m(est,sensor,fecha_ini, fecha_fin,folder):

    month = ["enero","febrero","marzo","abril","mayo","junio","julio","agosto","septiembre","octubre","noviembre","diciembre"]
    mes_n=fecha_ini.month
    mes_t=month[mes_n-1]
    folder_mes = os.path.dirname(os.path.abspath(__file__))+f"/Inf_mensuales/{est}_{sensor}/{fecha_ini.year}/{mes_t}/"
    if os.path.exists(folder_mes) == False:
        os.makedirs(folder_mes)

        for com in ["Z","N","E"]:

            try:
                
                list_est_sen = glob.glob(os.path.dirname(os.path.abspath(__file__))+f"/Datos/CM/Tablas_Cal_tiempos/{est}_{sensor}Z*")
                cod_sensor=[]
                if len(list_est_sen) == 2:
                    for s in list_est_sen:
                        if s[-10:-8] not in cod_sensor:
                            cod_sensor.append(s[-10:-8])
                    for cod_s in cod_sensor:
                        file = glob.glob(os.path.dirname(os.path.abspath(__file__))+f"/Datos/CM/Tablas_Cal_tiempos/{est}_{sensor}{com}_{cod_s}*")
                        
                        plt_disp = Plot(est,sensor, fecha_ini,fecha_fin,file)
                        plt_disp.plot_disp_fij(com,folder_mes,two_sensores=True)
                        plt_disp.plot_gapover_fij(com, folder_mes, two_sensores=True)
                        plt_disp.plot_off_fij(com, folder_mes,two_sensores=True)
                        plt_disp.plot_ppsd_fij(com, folder_mes,two_sensores=True)
                
                if len(list_est_sen) == 1:
                    
                    file = glob.glob(os.path.dirname(os.path.abspath(__file__))+f"/Datos/CM/Tablas_Cal_tiempos/{est}_{sensor}{com}*")
                    
                    plt_disp = Plot(est,sensor, fecha_ini,fecha_fin,file)
                    plt_disp.plot_disp_fij(com,folder_mes)
                    plt_disp.plot_gapover_fij(com, folder_mes)
                    plt_disp.plot_off_fij(com, folder_mes)
                    plt_disp.plot_ppsd_fij(com, folder_mes)
                    
            except:
                print("No hay componentes horizontales")
                continue
            
        hoy = datetime.now()
        fecha_hoy= f"{hoy.year}-{hoy.month}-{hoy.day}"
        #Creacion de Json
        datos_json = [{'funcionamiento':" ",'disponibilidad':" ",'gapover':" ", 'calidad':" ", 'offset':" ","ruido":" ","ultima_v":["","","",""],"recomendaciones":" ","fecha_creacion":f"{fecha_hoy} ","n_img_ruido":0,"n_img_recom":0}]
        
        list_est_sen = glob.glob(os.path.dirname(os.path.abspath(__file__))+f"/Datos/CM/Tablas_Cal_tiempos/{est}_{sensor}Z*")
        cod_sensor=[]
        
        if len(list_est_sen) == 2:
            for s in list_est_sen:
                if s[-10:-8] not in cod_sensor:
                    cod_sensor.append(s[-10:-8])
            for cod_s in cod_sensor:
                         
                with open(folder_mes + f"inf_{est}_{sensor}{cod_s}.json", 'w') as (file):
                    json.dump(datos_json, file, indent=4)
                    
        if len(list_est_sen) == 1:
            with open(folder_mes + f"inf_{est}_{sensor}.json", 'w') as (file):
                json.dump(datos_json, file, indent=4)
    

def actualizar_info_s(est,sensor,fecha_ini, fecha_fin,folder,semestre):
    

    folder_semestre = os.path.dirname(os.path.abspath(__file__))+f"/Inf_semestrales/{est}_{sensor}/{semestre}/"
    if os.path.exists(folder_semestre) == True:

        for com in ["Z","N","E"]:

            try:
                
                
                list_est_sen = glob.glob(os.path.dirname(os.path.abspath(__file__))+f"/Datos/CM/Tablas_Cal_tiempos/{est}_{sensor}Z*")
                cod_sensor=[]
                if len(list_est_sen) == 2:
                    for s in list_est_sen:
                        if s[-10:-8] not in cod_sensor:
                            cod_sensor.append(s[-10:-8])
                    for cod_s in cod_sensor:
                        file = glob.glob(os.path.dirname(os.path.abspath(__file__))+f"/Datos/CM/Tablas_Cal_tiempos/{est}_{sensor}{com}_{cod_s}*")
                        
                        plt_disp = Plot(est,sensor, fecha_ini,fecha_fin,file)
                        plt_disp.plot_disp_fij(com,folder_semestre,two_sensores=True)
                        plt_disp.plot_gapover_fij(com, folder_semestre, two_sensores=True)
                        plt_disp.plot_off_fij(com, folder_semestre,two_sensores=True)
                        plt_disp.plot_ppsd_fij(com, folder_semestre,two_sensores=True)
                
                if len(list_est_sen) == 1:
                    
                    file = glob.glob(os.path.dirname(os.path.abspath(__file__))+f"/Datos/CM/Tablas_Cal_tiempos/{est}_{sensor}{com}*")
                    
                    plt_disp = Plot(est,sensor, fecha_ini,fecha_fin,file)
                    plt_disp.plot_disp_fij(com,folder_semestre)
                    plt_disp.plot_gapover_fij(com, folder_semestre)
                    plt_disp.plot_off_fij(com, folder_semestre)
                    plt_disp.plot_ppsd_fij(com, folder_semestre)
                    
                
                #plt_disp.plot_disp_din(com,save=True, folder=folder_semestre) #Disponibilidad
                
            
            except:
                print("No hay componentes horizontales")
                continue

#Estado actual de la estación            
def estado_a(est,sensor="00"):
    
    file_esta = os.path.dirname(os.path.abspath(__file__))+f"/Estados/{est}.csv"


    if os.path.exists(file_esta) == True:        
    
        tab_est = pd.read_csv(file_esta)
        ultimo = tab_est.tail(1)
        f =ultimo["fecha"].iloc[0]
        act_sis = ultimo["act_sis"].iloc[0]
        act_elec =  ultimo["act_elec"].iloc[0]
        f_prob = ultimo["f_prob"].iloc[0]
        e_disp = ultimo['e_disp'].iloc[0]
        p_sist = ultimo["p_sist"].iloc[0]
        
        st.success(f"""
                **Observación Sismologo**\n
                {act_sis}\n
                **Observación Electrónico**\n
                {act_elec}\n
                **Fecha del problema:** {f_prob}\n
                
                **Estado de disponibilidad:**           {e_disp}\n
                **Problema de sistema:**           {p_sist}\n
                ______________
                Actualizado el {f}""")


    
        #fecha_man = fil_est3["FECHA MANTENIMIENTO"].iloc[0]
    
    else:
        st.write("Sin estado")

def actualizar_estado(est,sensor,act_sis,e_disp,p_sist,f_prob,f_act,act_elec):
    
    file_esta = os.path.dirname(os.path.abspath(__file__))+f"/Estados/{est}.csv"
    
    if os.path.exists(file_esta) == True:
        
        hoy = datetime.now()
        hoy = f_act
        text_fech_creacion= f"{hoy.year}/{hoy.month}/{hoy.day}"
        
        tab_est = pd.read_csv(file_esta)
        n_dict = pd.DataFrame({'fecha':[text_fech_creacion], 'act_sis':[act_sis],"act_elec":[act_elec],'e_disp':[e_disp],"p_sist":[p_sist],"f_prob":[f_prob]})
        tab_est2 = tab_est.append(n_dict,ignore_index=True)

        tab_est2.to_csv(file_esta, index=False)
        
    
    else:
        
        hoy = datetime.now()
        hoy = f_act
        text_fech_creacion= f"{hoy.day}/{hoy.month}/{hoy.year}"
        
        n_dict = {'fecha':[text_fech_creacion], 'act_sis':[act_sis],"act_elec":[act_elec],'e_disp':[e_disp],"p_sist":[p_sist],"f_prob":[f_prob]}
        df = pd.DataFrame(n_dict)
        df.to_csv(file_esta, index=False)

def reporte(resultado, ruta):

    st.sidebar.write("👇")
    # Guarda el reporte en una ruta especifica
    resultado.to_excel(os.path.dirname(os.path.abspath(__file__))+'/Reportes/' + ruta, index=False, startrow=7)
    libro = openpyxl.load_workbook(os.path.dirname(os.path.abspath(__file__))+'/Reportes/' + ruta)
    hoja = libro.active


    imagen = Image(os.path.dirname(os.path.abspath(__file__))+'/images/Plantilla.png')
    imagen.height = 66.6 #1.8#39.22%
    imagen.width = 170 #4.59
    
    hoja.add_image(imagen, 'A4')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    hoja['D4'] = "Informe del Estado Actual de las estaciones."
    hoja["D4"].alignment = Alignment(horizontal="center")
    hoja["D4"].font = Font(bold=True, size=20)
    hoja["D4"].alignment = Alignment(horizontal="center")

    fecha = datetime.now().strftime("%Y-%m-%d")
    hoja['L4'] = f'Fecha de creación: {fecha}'
    hoja["L4"].alignment = Alignment(horizontal="center")
    #hoja['A3'] = "______________________________"
    hoja.cell(row=8, column=1).border = thin_border
    hoja.merge_cells("A4:C7")
    hoja.merge_cells("D4:K7")
    hoja.merge_cells("L4:L7")
    

    # Cambia el color del fondo de los titulos
    for celda in hoja[8]:
        celda.fill = PatternFill(start_color='8CA448', end_color='8CA448', fill_type='solid')
                
    # Ajusta el anchos de las columnas para los datos del reporte
    for columna in hoja.columns:
        max_length = 23
        column = columna[0].column_letter
        
        len_cel = [len(str(celda.value)) for celda in columna[6:]] 
        max_length_col = np.max(len_cel)

        if max_length_col  < max_length:
            max_length_f = max_length_col+2
        if max_length_col  > max_length:
            max_length_f = 45

        
        adjusted_width = max_length_f
        hoja.column_dimensions[column].width = adjusted_width
    ##Agregar Colores en celda condicionados a sus valores en el excel
    #%Disponibilidad
    for celda,fila in zip(hoja["C"][8:],range(9,len(hoja["C"])+1)):

        color = "ffffffff"  #blanco color ARGB
        if celda.value  >= 80 and celda.value  <= 100:
            color = "ff92d050"  #verde color ARGB
        if celda.value  >= 60 and celda.value  < 80:
            color = "ff00b0f0"  #azul color ARGB
        if celda.value  >= 40 and celda.value  < 60:
            color = "ffffff00"  #amarillo color ARGB
        if celda.value  >= 20 and celda.value  < 40:
            color = "ffffc000"  #naranja color ARGB
        if celda.value  >= 0 and celda.value  < 20:
            color = "ffff0000" #rojo color ARGB

        hoja[f"C{fila}"].fill = PatternFill("solid", fgColor=color) ##Color de celda
    
    #Promedio de Gaps
    for celda,fila in zip(hoja["D"][8:],range(9,len(hoja["D"])+1)):
        color = "ffff0000" #rojo color ARGB
        if celda.value != None:
            if celda.value >= 9:
                color = "ffff0000" #rojo color ARGB
            if celda.value >= 3 and celda.value < 9:
                color = "ffffff00"  #amarillo color ARGB
            if celda.value < 3:
                color = "ff92d050"  #verde color ARGB

        if celda.value == None: 
            hoja[f"D{fila}"] = "None"
        hoja[f"D{fila}"].fill = PatternFill("solid", fgColor=color) ##Color de celda

    #Promedio de offset
    for celda,fila in zip(hoja["E"][8:],range(9,len(hoja["E"])+1)):
        color = "ffff0000" #rojo color ARGB
        if celda.value != None:
            if celda.value >= 50000 or celda.value <= -50000:
                color = "ffff0000" #rojo color ARGB
            if celda.value >= 10000 and celda.value < 50000  or celda.value <= -10000 and celda.value > -50000:
                color = "ffffff00"  #amarillo color ARGB
            if celda.value < 10000 and celda.value > -10000:
                color = "ff92d050"  #verde color ARGB

        if celda.value == None: 
            hoja[f"E{fila}"] = "None"
        hoja[f"E{fila}"].fill = PatternFill("solid", fgColor=color) ##Color de celda
    
    #Promedio de picos
    for celda,fila in zip(hoja["F"][8:],range(9,len(hoja["F"])+1)):
        color = "ffff0000" #rojo color ARGB
        if celda.value != None:
            if celda.value >= 10:
                color = "ffff0000" #rojo color ARGB
            if celda.value >= 6 and celda.value < 10:
                color = "ffffc000"  #naranja color ARGB
            if celda.value >= 1 and celda.value < 6:
                color = "ffffff00"  #amarillo color ARGB
            if  celda.value < 1:
                color = "ff92d050"  #verde color ARGB

        if celda.value == None: 
            hoja[f"F{fila}"] = "None"
        hoja[f"F{fila}"].fill = PatternFill("solid", fgColor=color) ##Color de celda
        
    
    #Promedio ppsd
    for celda,fila in zip(hoja["G"][8:],range(9,len(hoja["G"])+1)):
        color = "ffff0000" #rojo color ARGB
        if celda.value != None:
            if celda.value >= 90:
                color = "ffff0000" #rojo color ARGB
            if celda.value >= 70 and celda.value < 90:
                color = "ffffc000"  #naranja color ARGB
            if  celda.value < 70:
                color = "ff92d050"  #verde color ARGB

        if celda.value == None: 
            hoja[f"G{fila}"] = "None"
        hoja[f"G{fila}"].fill = PatternFill("solid", fgColor=color) ##Color de celda
    
    #Ultima visita
    for celda,fila in zip(hoja["H"][8:],range(9,len(hoja["H"])+1)):

        color = "ffff0000" #rojo color ARGB
        if celda.value != "sin datos":
            
            year_, month_, day_ = int(celda.value.split("-")[0]), int(celda.value.split("-")[1]), int(celda.value.split("-")[2])
            fecha_ = datetime(year_, month_,day_)
            
            year_ago=datetime.now()-timedelta(days=365)
            year_ago_10=datetime.now()-timedelta(days=300)
            if fecha_ <= year_ago:
                color = "ffff0000" #rojo color ARGB
            if fecha_ > year_ago and fecha_ <= year_ago_10:
                color = "ffffc000"  #naranja color ARGB
            if  fecha_ > year_ago_10:
                color = "ff92d050"  #verde color ARGB

        hoja[f"H{fila}"].fill = PatternFill("solid", fgColor=color) ##Color de celda
    
    #Estado de disponibilidad
    for celda,fila in zip(hoja["J"][8:],range(9,len(hoja["J"])+1)):

        if celda.value != None:
            if celda.value == "Bien":
                color = "ff92d050"  #verde color ARGB
            if celda.value == "Intermitente":
                color = "ffffc000"  #naranja color ARGB
            if celda.value == "Por fuera":
                color = "ffff0000" #rojo color ARGB
        
            hoja[f"J{fila}"].fill = PatternFill("solid", fgColor=color) ##Color de celda

            

    libro.save(os.path.dirname(os.path.abspath(__file__))+'/Reportes/' + ruta)
                
    # Permite descargar el reporte con el boton "Descargar Reporte"
    with open(os.path.dirname(os.path.abspath(__file__))+'/Reportes/' + ruta, 'rb') as file:
        contenido = file.read()
    st.sidebar.download_button(label='Descargar Reporte', data=BytesIO(contenido), file_name=ruta)

def create_db_all_state():

    #(self, estacion,sensor, fecha_ini,fecha_fin,list_est):

    #self.folder = download_folder
    #self.estacion = estacion
    #self.sensor = sensor
    #self.fecha_ini = fecha_ini
    #self.fecha_fin = fecha_fin
    #self.list_est = list_est
    

    list_est = sorted(glob.glob(os.path.dirname(os.path.abspath(__file__))+f"/Datos/CM/Tablas_Cal_/*"))
    #st.markdown("**Estacion   | Canal  | Sensor  | %disp** |Prom Gaps | Prom offset | Prom_picos | ppsd_prom | ultima visita |resp_tem | resp_elec")

    
    ##READ Ultima Visita
    file_xlsx = os.path.dirname(os.path.abspath(__file__))+"/dat_est/lastServiceMaintenance.xlsx"
    tab_est_uv = pd.read_excel(file_xlsx)

    #READ RESPONSABLES
    file_xlsx = os.path.dirname(os.path.abspath(__file__))+"/dat_est/stationReportCodLocation.xlsx"
        
    tab_est = pd.read_excel(file_xlsx)

    al_estacion, al_canal, al_sensor, al_prom_dis, al_gaps_prom, al_offset_prom, al_pic_prom, al_ppsd_prom, al_fecha_man, al_resp_tem, al_resp_elect = [],[],[],[],[],[],[],[],[],[],[]
    al_fecha_prob, al_estado_disp, al_probl_sist, al_act_sis,al_act_elec,al_fecha_act = [],[],[],[],[],[]
    l_estacion, l_canal, l_sensor, l_prom_dis, l_gaps_prom, l_offset_prom, l_pic_prom, l_ppsd_prom, l_fecha_man, l_resp_tem, l_resp_elect = [],[],[],[],[],[],[],[],[],[],[]
    l_fecha_prob, l_estado_disp, l_probl_sist, l_act_sis, l_act_elec, l_fecha_act = [],[],[],[],[],[]
    al_dict_dat = {}
    dict_dat = {}
    for all_est in list_est:
        
        #filtro por Estación
        estacion = all_est.split("/")[-1].split("_")[0]
        filtro_estacion = (tab_est["IDENTIFICADOR"] == estacion)
        fil_est = tab_est[filtro_estacion]
        
        if len(fil_est) > 0:

            if fil_est["ESTADO ESTACIÓN"].iloc[0] != "RETIRADA":       

                estacion = all_est.split("/")[-1].split("_")[0]
                canal = all_est.split("/")[-1].split("_")[1]
                sensor = all_est.split("/")[-1].split("_")[2]

                df_est = pd.read_csv(all_est)

                #Cambio de formato de fecha a datetime
                f_date = []
                for e in df_est["fecha"]:
                    year, month, day = e.split("-")
                    f_date.append(datetime(int(year),int(month),int(day)))
                
                #Cambio de formato de fecha times_gaps a datetime
                df_est["fecha"] = f_date

                #filtro por fechas seleccionadas
                
                ayer= datetime.now()-timedelta(days=3)
                mes_antes=datetime.now()-timedelta(days=18)
                filtro_fecha = (df_est["fecha"] >= mes_antes) & (df_est["fecha"] <= ayer) 
                fil_time = df_est[filtro_fecha]

                #DATOS
                #Disponibilidad
                prom_dis = round(fil_time["disponibilidad"].mean(),1)
                
                if len(fil_time["disponibilidad"]) == 0:
                    prom_dis = 0
                    
                
                #gaps
                # promedio de gaps y mas
                lista_gaps_sin_nueves=[]
                for i in range(0,len(fil_time["fecha"])):
                    if np.isnan(fil_time["num_gaps"].iloc[i]):
                        continue
                    if fil_time["num_gaps"].iloc[i] != -9: #se le puso -9 a los datos no disponibles
                        lista_gaps_sin_nueves.append(fil_time["num_gaps"].iloc[i])
                
                if len(lista_gaps_sin_nueves) !=0:
                    gaps_prom = round(np.mean(np.array(lista_gaps_sin_nueves)),2)
                    if gaps_prom == 0 and prom_dis == 0:
                        gaps_prom = None
                if len(lista_gaps_sin_nueves) ==0:
                    gaps_prom=None
                

                #offset
                ###Datos promedio, y mas
                lista_offset_sin_ceros=[]
                for of in fil_time["offs"]:
                    if str(of) != "nan" and of != None and of != -9 :
                        
                        lista_offset_sin_ceros.append(of)
                        
                
                
                if len(lista_offset_sin_ceros) >0:
                    offset_prom=round(np.mean(np.array(lista_offset_sin_ceros)),2)
                else:
                    offset_prom=None
                
                #Picos
                lista_picos_sin_nueves=[]
                
                for i in range(0,len(fil_time["fecha"])):
                    if np.isnan(fil_time["peaks"].iloc[i]):
                        continue
                    if fil_time["peaks"].iloc[i] != -9: #se le puso -9 a los datos no disponibles
                        lista_picos_sin_nueves.append(fil_time["peaks"].iloc[i]) 
                
                pic_prom=round(np.mean(np.array(lista_picos_sin_nueves)),2)
                if pic_prom == 0 and prom_dis == 0:
                    pic_prom = None

                ##PPSD
                lista_ppsd_sin_menosnueves=[]
                for of in fil_time["p_ppsd"]:
                    if np.isnan(of):
                        continue
                    if of != -9:
                        lista_ppsd_sin_menosnueves.append(of)

                if len(lista_ppsd_sin_menosnueves) != 0: 
                    ppsd_prom = round(np.mean(np.array(lista_ppsd_sin_menosnueves)),2)
                if len(lista_ppsd_sin_menosnueves) == 0: 
                    ppsd_prom = None

                ##Ultima Visita
                tab_est2 = tab_est_uv.drop(tab_est.index[[0]])
                tit=["ESTACION","ESTADO ESTACION","FECHA MANTENIMIENTO","TIPO DE MANTENIMIENTO", "RED DE MONITOREO","COMISIONADOS",	"REGISTRO CARGADO POR",	"COMENTARIOS"]
                tab_est2.columns = tit

                #filtro por Estación
                filtro_estacion = (tab_est2["ESTACION"] == estacion)
                fil_est3 = tab_est2[filtro_estacion]
                #st.dataframe(fil_est3)
                
                if len(fil_est3) > 0:
                    fecha_man = fil_est3["FECHA MANTENIMIENTO"].iloc[0]
                    tip_man=fil_est3["TIPO DE MANTENIMIENTO"].iloc[0].title()
                    vis_por = fil_est3["COMISIONADOS"].iloc[0].title()
                    comentarios = fil_est3["COMENTARIOS"].iloc[0].lower()
                else:
                    fecha_man = "sin datos"


                #RESPONSABLES
                
                #filtro por Estación
                filtro_estacion = (tab_est["IDENTIFICADOR"] == estacion)
                fil_est = tab_est[filtro_estacion]
                
                if len(fil_est) > 0:
                    
                    resp_elect1 =str(fil_est["ELECTRÓNICO RESPONSABLE"].iloc[0])
                    resp_tem1 = str(fil_est["TEMÁTICO RESPONSABLE"].iloc[0])
                    
                    resp_elect = resp_elect1.split(" ")[0]
                    resp_tem = resp_tem1.split(" ")[0]
                    
                    if len(resp_elect1.split(" ")) > 1:
                        resp_elect = resp_elect1.split(" ")[0]+" "+resp_elect1.split(" ")[1][0]+"."
                    if len(resp_tem1.split(" ")) > 1:
                        resp_tem = resp_tem1.split(" ")[0]+" "+resp_tem1.split(" ")[1][0]+"."
                        
                    
                    
                    
                    
                else:
                    resp_elect = "sin datos"
                    resp_tem = "sin datos"
                
                #fe = fil_time["fecha"].iloc[0]
                
                
                ##ESTADO
                file_esta = os.path.dirname(os.path.abspath(__file__))+f"/Estados/{estacion}.csv"

                if os.path.exists(file_esta) == True:        
                    
                    tab_est_estado = pd.read_csv(file_esta)
                    ultimo = tab_est_estado.tail(1)
                    fecha_act =ultimo["fecha"].iloc[0]
                    act_sis = ultimo["act_sis"].iloc[0]
                    act_elec =  ultimo["act_elec"].iloc[0]
                    f_prob = ultimo["f_prob"].iloc[0]
                    e_disp = ultimo['e_disp'].iloc[0]
                    p_sist = ultimo["p_sist"].iloc[0]
                
                else:
                    #f = None
                    act_sis = None
                    act_elec = None
                    f_prob = None
                    e_disp = None
                    p_sist = None  
                    fecha_act = None
                
                        
                #st.write(estacion,"      |",canal,"       |",sensor,"       |",str(round(prom_dis,2)),"   |", str(gaps_prom),"    |", str(offset_prom)," |",str(pic_prom),"|", str(ppsd_prom),"|", str(fecha_man)
                #         ,"|",resp_tem,"|",resp_elect)


                al_estacion.append(estacion)
                al_canal.append(canal)
                al_sensor.append(sensor)
                al_prom_dis.append(round(prom_dis,2))
                al_gaps_prom.append(gaps_prom)
                al_offset_prom.append(offset_prom)
                al_pic_prom.append(pic_prom)
                al_ppsd_prom.append(ppsd_prom)
                al_fecha_man.append(fecha_man)
                al_resp_tem.append(resp_tem)
                al_resp_elect.append(resp_elect)
                #Estado
                al_fecha_prob.append(f_prob)
                al_estado_disp.append(e_disp)
                al_probl_sist.append(p_sist)
                al_act_sis.append(act_sis)
                al_act_elec.append(act_elec)
                al_fecha_act.append(fecha_act)
                
                if canal[-1] == "Z":
                    l_estacion.append(estacion)
                    l_canal.append(canal)
                    l_sensor.append(sensor)
                    l_prom_dis.append(round(prom_dis,1))
                    l_gaps_prom.append(gaps_prom)
                    l_offset_prom.append(offset_prom)
                    l_pic_prom.append(pic_prom)
                    l_ppsd_prom.append(ppsd_prom)
                    l_fecha_man.append(fecha_man)   
                    l_resp_tem.append(resp_tem)
                    l_resp_elect.append(resp_elect)
                    #Estado
                    l_fecha_prob.append(f_prob)
                    l_estado_disp.append(e_disp)
                    l_probl_sist.append(p_sist)
                    l_act_sis.append(act_sis)
                    l_act_elec.append(act_elec)
                    l_fecha_act.append(fecha_act)
        
    
    #"**Estacion   | Canal  | Sensor  | %disp** |Prom Gaps | Prom offset | Prom_picos | ppsd_prom | ultima visita |resp_tem | resp_elec"
    al_dict_dat["Est."] = al_estacion
    al_dict_dat["can_"] = al_canal
    al_dict_dat["Sen."] = al_sensor
    al_dict_dat["%disp."] = al_prom_dis
    al_dict_dat["p_gaps"] = al_gaps_prom
    al_dict_dat["p_offset"] = al_offset_prom
    al_dict_dat["p_picos"] = al_pic_prom
    al_dict_dat["ppsd_p"] = al_ppsd_prom
    al_dict_dat["Ult_visita"] = al_fecha_man
    al_dict_dat["Sismologo_Responsable"] = al_resp_tem
    al_dict_dat["Electronico_responsable"] = al_resp_elect
    #Estado
    al_dict_dat["Fecha_problema"] = al_fecha_prob
    al_dict_dat["Estado_disp."] = al_estado_disp
    al_dict_dat["Estado_sistema"] = al_probl_sist
    al_dict_dat["Obs._sismologo"] = al_act_sis
    al_dict_dat["Obs._electronico"] = al_act_elec
    al_dict_dat["Fecha_actualizacion"] = al_fecha_act
    
    ###
    dict_dat["Est."] = l_estacion
    dict_dat["can_"] = l_canal
    dict_dat["Sen."] = l_sensor
    dict_dat["%disp."] = l_prom_dis
    dict_dat["p_gaps"] = l_gaps_prom
    dict_dat["p_offset"] = l_offset_prom
    dict_dat["p_picos"] = l_pic_prom
    dict_dat["ppsd_p"] = l_ppsd_prom
    dict_dat["Ult_visita"] = l_fecha_man
    dict_dat["Sismologo_Responsable"] = l_resp_tem
    dict_dat["Electronico_responsable"] = l_resp_elect
    #Estado
    dict_dat["Fecha_problema"] = l_fecha_prob
    dict_dat["Estado_disp."] = l_estado_disp
    dict_dat["Estado_sistema"] = l_probl_sist
    dict_dat["Obs._sismologo"] = l_act_sis
    dict_dat["Obs._electronico"] = l_act_elec
    dict_dat["Fecha_actualizacion"] = l_fecha_act
    
    df_estado = pd.DataFrame(dict_dat)
    
    df_estado = df_estado.reindex(columns=["Est.", "Sen.", "%disp.", "p_gaps", "p_offset", "p_picos", "ppsd_p",
                                                 "Ult_visita", "Fecha_problema","Estado_disp.","Estado_sistema","Obs._sismologo","Obs._electronico","Sismologo_Responsable","Electronico_responsable","Fecha_actualizacion"])
    
    df_estado = df_estado.sort_values("Estado_disp.",ascending=False)
    #st.write(df_estado)
    df_al_estado = pd.DataFrame(al_dict_dat)

    ###Filtro por persona responsable
    #df_estado

    l_resp_sis = []
    for r_sis, r_elec in zip(df_estado["Sismologo_Responsable"],df_estado["Electronico_responsable"]):
        if r_sis not in l_resp_sis:
            l_resp_sis.append(r_sis)
        if r_elec not in l_resp_sis:
            l_resp_sis.append(r_elec)
    l_resp_sis = sorted(l_resp_sis)

    col1, col2 = st.columns([1,3])
    with col1:
        ##Selección de Responsable
        option_resp = st.selectbox(
        "**Responsable**",
        l_resp_sis,
        index=None,
        placeholder="Select contact method...",)
    with col2:
        st.write("")
    
    if option_resp in list(df_estado["Sismologo_Responsable"]):
        filtro_resp = (df_estado["Sismologo_Responsable"] == option_resp)  
        df_estado = df_estado[filtro_resp]
    
    if option_resp in list(df_estado["Electronico_responsable"]):
        filtro_resp = (df_estado["Electronico_responsable"] == option_resp)  
        df_estado = df_estado[filtro_resp]
    ####
    
    #LISTAS DE COLORES POR COLUMNA
    l_col_estad,l_col_disp,l_col_gaps,l_col_offset,l_col_pic,l_col_ppsd,l_col_fman = [],[],[],[],[],[],[]
    
    ### estado disponibilidad
    for d in range(len(list(df_estado["Est."]))):
        #if d <= len(l_estado_disp)-1:
            
        if df_estado["Estado_disp."].iloc[d] == "Bien":
            color = '#92d050'
        if df_estado["Estado_disp."].iloc[d] == "Intermitente":
            color = "#ffc000" 
        if df_estado["Estado_disp."].iloc[d] == "Por fuera":
            color = 'red'
        if df_estado["Estado_disp."].iloc[d] == None:
            color = "white"
                
        l_col_estad.append(f'background-color: {color}')
    
    ###promedio de disponibilidad  
    for dis in list(df_estado["%disp."]):
        color = "white"
        if dis >= 80 and dis <= 100:
            color = "#92d050"  #verde
        if dis >= 60 and dis < 80:
            color = "#00b0f0"  #azul
        if dis >= 40 and dis < 60:
            color = "#ffff00"  #amarillo
        if dis >= 20 and dis < 40:
            color = "#ffc000"  #naranja
        if dis >= 0 and dis < 20:
            color = "red"
        
        l_col_disp.append(f'background-color: {color}')
    
    #promedio de gaps
    for gaps_ in list(df_estado["p_gaps"]):
        color = "red"
        if gaps_ >= 9:
            color = "red"
        if gaps_ >= 3 and gaps_ < 9:
            color = "#ffff00"
        if gaps_ < 3:
            color = "#92d050"

        l_col_gaps.append(f'background-color: {color}')

    #promedio de offset
    for off__ in list(df_estado["p_offset"]):
        color_off = "red"
        if off__ >= 50000 or off__ <= -50000:
            color_off = "red"
        if off__ >= 10000 and off__ < 50000  or off__ <= -10000 and off__ > -50000:
            color_off = "#ffff00"
        if off__ < 10000 and off__ > -10000:
            color_off = "#92d050"

        l_col_offset.append(f'background-color: {color_off}')
    
    #promedio de picos
    for pic__ in list(df_estado["p_picos"]):

        color_pic = "red"
        if pic__ >= 10:
            color_pic = "red"
        if pic__ >= 6 and pic__ < 10:
            color_pic = "#ffc000"
        if pic__ >= 1 and pic__ < 6:
            color_pic = "#ffff00"
        
        if  pic__ < 1:
            color_pic = "#92d050"

        l_col_pic.append(f'background-color: {color_pic}')
    
    
    #promedio de ppsd
    for ppsd__ in list(df_estado["ppsd_p"]):
        
        if ppsd__ != None:
            color_ppsd = "red"
            if ppsd__ >= 90:
                color_ppsd = "red"
            if ppsd__ >= 70 and ppsd__ < 90:
                color_ppsd = "#ffc000"
            if  ppsd__ < 70:
                color_ppsd = "#92d050"

            l_col_ppsd.append(f'background-color: {color_ppsd}')
        else:
            l_col_ppsd.append(f'background-color: white')
    
    #Ultima visita
    for f_man__ in list(df_estado["Ult_visita"]):
        
        
        if f_man__ != "sin datos":
            
            year_, month_, day_ = int(f_man__.split("-")[0]), int(f_man__.split("-")[1]), int(f_man__.split("-")[2])
            fecha_ = datetime(year_, month_,day_)
            
            year_ago=datetime.now()-timedelta(days=365)
            year_ago_10=datetime.now()-timedelta(days=300)
            if fecha_ <= year_ago:
                color_f_u = "red"
            if fecha_ > year_ago and fecha_ <= year_ago_10:
                color_f_u = "#ffc000"
            if  fecha_ > year_ago_10:
                color_f_u = "#92d050"

            l_col_fman.append(f'background-color: {color_f_u}')
        else:
            l_col_fman.append(f'background-color: white')
          
        
        
    
   

    def run_colors(s):


        color = "white"
        nonlocal l_col_estad
        nonlocal l_col_disp
        nonlocal l_col_gaps
        nonlocal l_col_offset
        nonlocal l_col_pic
        nonlocal l_col_ppsd
        nonlocal l_col_fman

        if s.name == "Est." or s.name == "can_" or s.name == "Sen.":
            
            return ['background-color: white']*len(s)
        
        if s.name == "%disp.":
            return l_col_disp
        
        if s.name == "Estado_disp.":
            return l_col_estad

        if s.name == "p_gaps":
            return l_col_gaps
        
        if s.name == "p_offset":
            return l_col_offset
        
        if s.name == "p_picos":
            return l_col_pic
        
        if s.name == "ppsd_p":
            return l_col_ppsd
        
        if s.name == "Ult_visita":
            return l_col_fman
        
        return [f'background-color: {color}']*len(s)


    
    
    ##Ploteo de tabla de estados    
    
    st.dataframe(df_estado.style.apply(run_colors,axis=0).format(precision=2),  height=500)

    est1 = []
    for est in l_estacion:

        if est not in est1:
            est1.append(est)
            
    
    st.write("________________")
    c1, c2 = st.columns([0.8,3])
    with c1:
        option = st.selectbox('**Estación**',sorted(est1) ,key ="est_all")  
    with c2:
        st.write("")
    
    filtro_est = (df_al_estado["Est."] == option)  
    df_al_estado_filt = df_al_estado[filtro_est]
    df_al_estado_filt = df_al_estado_filt.drop(["Sismologo_Responsable","Electronico_responsable","Ult_visita","Fecha_problema","Estado_disp.",
                                                "Estado_sistema","Obs._sismologo","Obs._electronico"], axis=1)
    
    
    
    col1, col2 =st.columns([1,1.5])
    with col1:

        opt_f_act = datetime.now()
        f_prob = datetime.now()
        ##Comentario anterior
        file_esta = os.path.dirname(os.path.abspath(__file__))+f"/Estados/{option}.csv"
        if os.path.exists(file_esta) == True:        
        
            tab_est = pd.read_csv(file_esta)
            ultimo = tab_est.tail(1)
            #opt_f_act =ultimo["fecha"].iloc[0]
            #opt_f_act =datetime(int(ultimo["fecha"].iloc[0].split("-")[0]),
            #                    int(ultimo["fecha"].iloc[0].split("-")[1]),
            #                    int(ultimo["fecha"].iloc[0].split("-")[2]))
            act_sis = ultimo["act_sis"].iloc[0]
            act_elec =  ultimo["act_elec"].iloc[0]
            #f_prob = ultimo["f_prob"].iloc[0]
            f_prob =datetime(int(ultimo["f_prob"].iloc[0].split("-")[0]),
                                    int(ultimo["f_prob"].iloc[0].split("-")[1]),
                                    int(ultimo["f_prob"].iloc[0].split("-")[2]))
            #e_disp = ultimo['e_disp'].iloc[0]
            #p_sist = ultimo["p_sist"].iloc[0]

        if os.path.exists(file_esta) == False:    
            act_sis = ""
            act_elec = ""

        st.dataframe(df_al_estado_filt)
        

        st.write("________________")
        st.subheader("ACTUALIZAR ESTADO")
        #opt_f_act= st.date_input("**Fecha de actualizacion**", value=None)
        
        co1, co2,co3 = st.columns(3)
        with co1:  opt_e_disp = st.selectbox("**Estado Disponibilidad**", ("Bien","Intermitente","Por fuera"))
        with co2:  opt_p_sist=st.selectbox("**Problema sistema**", ("Instrumentación","Energía","Comunicaciones","Otro","ninguno"))
        with co3:  opt_f_prob= st.date_input("**Fecha del problema**", value=f_prob)
        
        #Observación del
        c1,c2,c3,c4 = st.columns(4)
        with c1:
            st.write(" ")
            st.write(" ")
            st.markdown("**Observación del**")
        with c2:
            opt_encar = st.selectbox(" ", ("Sismólogo","Electrónico",))
        with c3: st.write(" ")
        with c4: st.write(" ")
        #act_elec
        act_encar=st.text_area("",key="t_act")
        act_encar = act_encar.replace("\n"," ")
        
        ##
        
        if opt_encar == "Sismólogo":
            act_sis = act_encar
        if opt_encar == "Electrónico":
            act_elec = act_encar
        if opt_f_prob != None:
            f_prob = opt_f_prob
        
        if st.button("Actualizar estado", key="but_create"):
            actualizar_estado(option,"00",act_sis,opt_e_disp,opt_p_sist,f_prob,opt_f_act,act_elec) 
            st.rerun() 
    
    with col2:
        st.subheader("ESTADO")
        estado_a(option)
        
        st.write("_______")
        st.subheader("BITACORA")
        #BITACORA     
        ayer= datetime.now()-timedelta(days=3)
        mes_antes=datetime.now()-timedelta(days=95)
        
        get_bitacora(option, mes_antes, ayer)

        col1, col2 = st.columns([3,0.8])
        with col2:

            
            if st.button('Actualizar Bitacora',help="recargar y actualizar los datos con la información del SIIGEO"):
                actualizar_bitacora_siigeo()   
                st.rerun()
        with col1:
            st.markdown("**Esta información es extraida del SIIGEO**")

    
    if st.sidebar.button("Crear Documento", type="primary", key="but_create_ex"):
        
        if option_resp is None:
            option_resp = "todo"
        ruta = f'Estado_Actual_{option_resp}.xlsx'
        reporte(df_estado,ruta)
        
    

#def act_est_all_can():
    
    


#def tabla_estado():

