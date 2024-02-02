import streamlit as st
import os.path
import pandas as pd
#from Reporte import reporte
import openpyxl
import datetime
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.drawing.image import Image
from io import BytesIO



### Generacion y descarga del reporte
def reporte(resultado, ruta):
    # Guarda el reporte en una ruta especifica
    resultado.to_excel(os.path.dirname(os.path.abspath(__file__))+'/Reportes/' + ruta, index=False, startrow=9)
    libro = openpyxl.load_workbook(os.path.dirname(os.path.abspath(__file__))+'/Reportes/' + ruta)
    hoja = libro.active

    imagen = Image(os.path.dirname(os.path.abspath(__file__))+'/Imagenes/Plantilla.png')
    imagen.height = 140
    imagen.width = 300
    hoja.add_image(imagen, 'A1')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    fecha = datetime.datetime.now()
    hoja['A8'] = 'Fecha de creción:'
    hoja['B8'] = fecha
    hoja.cell(row=8, column=1).border = thin_border

    # Cambia el color del fondo de los titulos
    for celda in hoja[10]:
        celda.fill = PatternFill(start_color='8CA448', end_color='8CA448', fill_type='solid')
                
    # Ajusta el anchos de las columnas para los datos del reporte
    for columna in hoja.columns:
        max_length = 0
        column = columna[0].column_letter
        for celda in columna:
            try:
                if len(str(celda.value)) > max_length:
                    max_length = len(celda.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        hoja.column_dimensions[column].width = adjusted_width
    libro.save(os.path.dirname(os.path.abspath(__file__))+'/Reportes/' + ruta)
                
    # Permite descargar el reporte con el boton "Descargar Reporte"
    with open(os.path.dirname(os.path.abspath(__file__))+'/Reportes/' + ruta, 'rb') as file:
        contenido = file.read()
    st.download_button(label='Descargar Reporte', data=BytesIO(contenido), file_name=ruta)

    
# Funcion que se encarga de verificar los datos de las estaciones de la red de monitoreo
def verificar_uno(df_uno):
    st.dataframe(df_uno, hide_index=True, use_container_width=True)

    st.divider()

    st.subheader('Codigos de error', help='Desplegable que sirve de guía para conocer que errores y hallazgos busca el aplicativo en el reporte.')
    with st.expander('Errores y hallazgos:'):

        st.subheader('Errores:')

        st.info('1. El estado de la estación es "Retirada", pero no tiene ninguna fecha de retiro.')
        st.info('2. La columna hibrida es "Si" a la vez que la red de monitoreo no es "HIBRIDA".')
        st.info('3. La estación tiene varios codigos de localización, ejemplo "00 - 10", pero su red de monitoreo no es "HIBRIDA".')

        st.subheader('Hallazgos:')

        st.info('1. La columna “IDENTIFICADOR” está vacía.')
        st.info('2. La columna “NOMBRE” está vacía.')
        st.info('3. La columna “LATITUD (°)” está vacía.')
        st.info('4. La columna “LONGITUD (°)” está vacía.')
        st.info('5. La columna “ELEVACION (msnm)” está vacía.')
        st.info('6. La columna “DEPARTAMENTO” está vacía.')
        st.info('7. La columna “MUNICIPIO” está vacía.')
        st.info('8. La columna “ESTADO” está vacía.')
        st.info('9. La columna “RED MONITOREO” está vacía.')
        st.info('10. La columna “HIBRIDA” está vacía.')
        st.info('11. La columna “CODIGO DE LOCALIZACION” está vacía.')
        st.info('12. La columna “SUBRED” está vacía.')
        st.info('13. La columna “AGENCIA” está vacía.')
        st.info('14. La columna “FECHA INSTALACION” está vacía.')
        st.info('15. La columna “SISTEMA DE ENERGIA” está vacía.')
        st.info('16. La columna “CALIDAD DEL TERRENO” está vacía.')
        st.info('17. La columna “GEOLOGIA” está vacía.')
        st.info('18. La columna “TOPOGRAFIA” está vacía.')
        st.info('19. La columna “RESPONSABLE TEMATICO” está vacía.')
        st.info('20. La columna “RESPONSABLE TECNICO” está vacía.')
        st.info('21. La columna “ISOLUCION” está vacía.')
        st.info('22. La columna “ACCESO” está vacía.')
        st.info('23. El estado de la estación es "Activa", pero tiene una fecha de retiro.')

    st.divider()

    errores = {}
    hallazgos = {}

    # Selección de las estaciones a verificar
    st.subheader('Estaciones por verificar', help='Aquí se genera una vista previa del reporte con las incoherencias y errores de cada estación, el usuario puede elegir las estaciones que quiere verificar, en caso de no elegir ninguna el aplicativo verificara todas las estaciones disponibles.')
    def eleccion(datos):
        opcion = st.multiselect(
            'Seleccione las estaciones a verificar:',
            options=datos['IDENTIFICADOR'].unique()
        , key="mult_select_est0")

        if len(opcion) == 0:
            respuesta = datos
            nada = True
        else:
            respuesta = datos[datos['IDENTIFICADOR'].isin(opcion)]
            nada = False
        
        return respuesta, nada
    respuesta_uno, nada = eleccion(df_uno)

    # Seleccion de la agencia a la que pertenecen las estaciones a verificar
    st.subheader('Agencia', help='Si no se selecciona ninguna agencia, el aplicativo tomará las estaciones que tengan como agencia "SERVICIO GEOLOGICO COLOMBIANO"')
    def agencia(respuesta_uno):
        opciones = st.multiselect(
            'Seleccione las agencias a verificar:',
            options=respuesta_uno['AGENCIA'].unique(),
            default=['SERVICIO GEOLOGICO COLOMBIANO']
        , key="mult_select_agen1")

        if len(opciones) == 0:
            respuesta_dos = respuesta_uno
        else:
            respuesta_dos = respuesta_uno[respuesta_uno['AGENCIA'].isin(opciones)]
        
        return respuesta_dos
    respuesta_dos = agencia(respuesta_uno)
    respuesta = respuesta_dos.drop(['FECHA RETIRO'], axis=1)

    # Identificador de errores y hallazgos
    for row_index, row in respuesta_dos.iterrows():
        identificador = row['IDENTIFICADOR']

        # Error 1 "El estado de la estación es "Retirada", pero no tiene ninguna fecha de retiro."
        if (row['ESTADO'] == 'RETIRADA') and not(pd.notnull(row['FECHA RETIRO'])):
            if (row['ESTADO'] != 'PROXIMA A INSTALAR'):
                if identificador in errores:
                    errores[identificador] += ', Error 1'
                else:
                    errores[identificador] = 'Error 1'
        
        # Error 2 "La columna hibrida es "Si" a la vez que la red de monitoreo es diferente a "HIBRIDA"."
        if (row['HIBRIDA'] == 'Si') and (row['RED MONITOREO'] != 'HIBRIDA'):
            if (row['ESTADO'] != 'PROXIMA A INSTALAR'):
                if identificador in errores:
                    errores[identificador] += ', Error 2'
                else:
                    errores[identificador] = 'Error 2'
        
        # Verificamos que la estación cuente con más de un tipo diferente de sensor (sismometro, acelerografo, etc...)
        for codigos in row['CODIGO DE LOCALIZACION']:
            numeros = [int(num) for num in codigos.replace(' ', '').split('-') if num.isdigit()]
            if len(numeros) > 1:
                primer_numero = numeros[0]
                for numero in numeros[1:]:
                    if abs(primer_numero - numero) > 9:
                        # Error 3 "La estación tiene varios codigos de localización "00 - 10", pero su red de monitoreo no es "HIBRIDA"."
                        if ('-' in row['CODIGO DE LOCALIZACION']) and (row['RED MONITOREO'] != 'HIBRIDA'):
                            if (row['ESTADO'] != 'PROXIMA A INSTALAR'):
                                if identificador in errores:
                                    errores[identificador] += ', Error 3'
                                else:
                                    errores[identificador] = 'Error 3'
            else:
                pass
        
        if not(identificador in errores):
            errores[identificador] = 'Sin errores'

        # Hallazgo 1 "El estado de la estación es "Activa", pero tiene una fecha de retiro."
        #if (row['ESTADO'] == 'ACTIVA') and (pd.notnull(row['FECHA RETIRO'])):
        if (row['ESTADO'] == 'ACTIVA') and (row['FECHA RETIRO'] != 'None'):
            if (row['ESTADO'] != 'PROXIMA A INSTALAR'):
                if identificador in hallazgos:
                    hallazgos[identificador] += ', Hallazgo 23'
                else:
                    hallazgos[identificador] = 'Hallazgo 23'

        for columna in respuesta.columns:
            if (row['ESTADO'] != 'PROXIMA A INSTALAR'):
                if not pd.notnull(row[columna]):
                    # Asignar un número de hallazgo basado en la columna vacía
                    numero_hallazgo = respuesta.columns.get_loc(columna) + 1
                    
                    if (numero_hallazgo == 19) and ((row['ESTADO'] == 'RETIRADA') or (row['ESTADO'] == 'INACTIVA')):
                        pass
                    elif (numero_hallazgo == 20) and ((row['ESTADO'] == 'RETIRADA') or (row['ESTADO'] == 'INACTIVA')):
                        pass
                    else:
                        # Agregar el hallazgo al identificador correspondiente en el diccionario
                        if identificador in hallazgos:
                            hallazgos[identificador] += f', Hallazgo {numero_hallazgo}'
                        else:
                            hallazgos[identificador] = f'Hallazgo {numero_hallazgo}'
        
        if not(identificador in hallazgos):
            hallazgos[identificador] = 'Sin hallazgos'

    # Las listas se convierten en dataframes para poder unirlos con un merge
    errores_df = pd.DataFrame(list(errores.items()), columns=['ESTACION', 'ERRORES'])
    hallazgos_df = pd.DataFrame(list(hallazgos.items()), columns=['ESTACION', 'HALLAZGOS'])

    estado = respuesta_dos[['IDENTIFICADOR','ESTADO']]
    estado = estado.rename(columns={'IDENTIFICADOR':'ESTACION'})

    # Se unen los datos almacenados en los dataframes siempre y cuando compartan una misma estacion
    df_final_uno = pd.DataFrame()
    df_final_uno = estado.merge(
        errores_df,
        how='outer',
        on='ESTACION'
    )

    df_final = df_final_uno.merge(
        hallazgos_df,
        how='outer',
        on='ESTACION'
    )

    if nada is True:
        for row_index, row in df_final.iterrows():
            if row['ERRORES'] == 'Sin errores' and row['HALLAZGOS'] == 'Sin hallazgos':
                df_final = df_final.drop(index=row_index)
        st.dataframe(df_final, hide_index=True, use_container_width=True)
    elif df_final.empty:
        st.info('No se encontro ningún error o hallazgo')
    else:
        st.dataframe(df_final, hide_index=True, use_container_width=True)

    if len(df_final) != 0:
        ruta = 'Verificador_General.xlsx'
        reporte(df_final, ruta)
    else:
        st.success('No se encontraron ni errores ni hallazgos')

# Funcion que se encarga de verificar los datos de los instrumentos de las estaciones
def verificar_dos(df_dos):
    df_dos['CODIGO LOCALIZACION'] = df_dos['CODIGO LOCALIZACION'].replace(
        {
            0: '00',
            10: '10',
            11: '11',
            20: '20',
            30: '30',
            40: '40'
        }
    )
    
    st.dataframe(df_dos, hide_index=True, use_container_width=True)

    st.divider()

    errores = {}
    hallazgos = {}
    estado = {}
    codigo_localizacion = {}

    st.subheader('Codigos de error', help='El desplegable "Errores y hallazgos" contiene todas las comparaciones que evaluará el aplicativo en el reporte')
    with st.expander('Errores y hallazgos'):

        st.subheader('Errores:')

        st.info('1. El tipo de adquisición de la estación esta en "Tiempo Real", pero también tiene un tipo de almacenamiento.')
        st.info('2. El tipo de adquisición de la estación esta en "Descarga", pero el tipo de almacenamiento no tiene ningún dato.')
        st.info('3. El tipo de adquisición de la estación esta en "Descarga", pero también tiene un tipo de transmisión.')
        st.info('4. El estado de la estación esta en "RETIRADA" y tanto el sensor como el digitalizador tiene algún dato.')

        st.subheader('Hallazgos:')

        st.info('1. La columna "ID ESTACION" esta vacía.')
        st.info('2. La columna "NOMBRE" esta vacía.')
        st.info('3. La columna "LATITUD" esta vacía.')
        st.info('4. La columna "LONGITUD" esta vacía.')
        st.info('5. La columna "ELEVACION" esta vacía.')
        st.info('6. La columna "MUNICIPIO" esta vacía.')
        st.info('7. La columna "DEPARTAMENTO" esta vacía.')
        st.info('8. La columna "RED MONITOREO" esta vacía.')
        st.info('9. La columna "ESTADO ESTACION" esta vacía.')
        st.info('10. La columna "HIBRIDA" esta vacía.')
        st.info('11. La columna "ELECTRONICO RESPONSABLE" esta vacía.')
        st.info('12. La columna "TEMATICO RESPONSABLE" esta vacía.')
        st.info('13. La columna "CODIGO LOCALIZACION" esta vacía.')
        st.info('14. La columna "INTRUMENTO" esta vacía.')
        st.info('15. La columna "SERIAL SENSOR" esta vacía.')
        st.info('16. La columna "MARCA SENSOR" esta vacía.')
        st.info('17. La columna "MODELO SENSOR" esta vacía.')
        st.info('18. La columna "SERIAL DIGITALIZADOR" esta vacía.')
        st.info('19. La columna "MARCA DIGITALIZADOR" esta vacía.')
        st.info('20. La columna "MODELO DIGITALIZADOR" esta vacía.')
        st.info('21. La columna "FECHA INICIO" esta vacía.')
        st.info('22. La columna "FECHA FIN" esta vacía.')
        st.info('23. La columna "ESTADO CODIGO" esta vacía.')
        st.info('24. La columna "TIPO ALMACENAMIENTO" esta vacía.')
        st.info('25. La columna "CONDICION DE INSTALACION" esta vacía.')
        st.info('26. La columna "TIPO DE TRANSMISION" esta vacía.')
        st.info('27. La columna "TIPO DE DESCARGA" esta vacía.')
        st.info('28. La columna "TIPO DE ALCANCE" esta vacía.')
        st.info('29. La columna "TIPO DE ESTACION" esta vacía.')
        st.info('30. La columna "TIPO DE ADQUISICION" esta vacía.')
        st.info('31. La columna "COMENTARIOS" esta vacía.')
    
    st.divider()

    # Selección de las estaciones a verificar
    st.subheader('Estaciones por verificar', help='Aquí se genera una vista previa del reporte con las incoherencias y errores de cada estación, el usuario puede elegir las estaciones que quiere verificar, en caso de no elegir ninguna el aplicativo verificara todas las estaciones disponibles.')
    def eleccion(datos):
        df_filtrado = datos
        df_filtrado = df_filtrado.sort_values(by=['FECHA INICIO'], ascending=False)
        df_filtrado = df_filtrado.sort_values(by=['ID ESTACION'])

        opcion = st.multiselect(
            'Seleccione las estaciones a verificar:',
            options=df_filtrado['ID ESTACION'].unique()
        , key="mult_select_est1")

        if len(opcion) == 0:
            respuesta = df_filtrado
            nada = True
        else:
            respuesta = df_filtrado[df_filtrado['ID ESTACION'].isin(opcion)]
            nada = False
        
        return respuesta, nada
    respuesta, nada = eleccion(df_dos)

    respuesta = pd.DataFrame(respuesta)
    for row_index, row in respuesta.iterrows():
        identificador = row['ID ESTACION']
        codigo = row['CODIGO LOCALIZACION']

        estado[f'{identificador}({str(codigo).split(".", 1)[0]})'] = row['ESTADO ESTACION']
        codigo_localizacion[f'{identificador}({str(codigo).split(".", 1)[0]})'] = codigo

        # Error 1 "El tipo de adquisición de la estación esta en "Tiempo Real", pero también tiene un tipo de almacenamiento."
        if (row['TIPO DE ADQUISICION'] == 'Tiempo Real') and (pd.notnull(row['TIPO ALMACENAMIENTO'])):
            if identificador in errores:
                errores[f'{identificador}({str(codigo).split(".", 1)[0]})'] += ', Error 1'
            else:
                errores[f'{identificador}({str(codigo).split(".", 1)[0]})'] = 'Error 1'
        
        # Error 2 "El tipo de adquisición de la estación esta en "Descarga", pero el tipo de almacenamiento no tiene ningún dato."
        if (row['TIPO DE ADQUISICION'] == 'Descarga') and not(pd.notnull(row['TIPO ALMACENAMIENTO'])):
            if (f'{identificador}({str(codigo).split(".", 1)[0]})') in errores:
                errores[f'{identificador}({str(codigo).split(".", 1)[0]})'] += ', Error 2'
            else:
                errores[f'{identificador}({str(codigo).split(".", 1)[0]})'] = 'Error 2'
        
        # Error 3 "El tipo de adquisición de la estación esta en "Descarga", pero también tiene un tipo de transmisión."
        if (row['TIPO DE ADQUISICION'] == 'Descarga') and (pd.notnull(row['TIPO DE TRANSMISION'])):
            if (f'{identificador}({str(codigo).split(".", 1)[0]})') in errores:
                errores[f'{identificador}({str(codigo).split(".", 1)[0]})'] += ', Error 3'
            else:
                errores[f'{identificador}({str(codigo).split(".", 1)[0]})'] = 'Error 3'
        
        # Error 4 "El estado de la estación esta en "RETIRADA" y tanto el sensor como el digitalizador tiene algún dato."
        if (row['ESTADO ESTACION'] == 'RETIRADA') and (pd.notnull(row['SERIAL SENSOR']) or pd.notnull(row['SERIAL DIGITALIZADOR'])):
            if (f'{identificador}({str(codigo).split(".", 1)[0]})') in errores:
                errores[f'{identificador}({str(codigo).split(".", 1)[0]})'] += ', Error 4'
            else:
                errores[f'{identificador}({str(codigo).split(".", 1)[0]})'] = 'Error 4'
        
        if not((f'{identificador}({str(codigo).split(".", 1)[0]})') in errores):
            errores[f'{identificador}({str(codigo).split(".", 1)[0]})'] = 'Sin errores'

        for columna in respuesta.columns:
            if (row['ESTADO ESTACION'] != 'PROXIMA A INSTALAR'):
                if not pd.notnull(row[columna]):
                    # Asignar un número de hallazgo basado en la columna vacía
                    numero_hallazgo = respuesta.columns.get_loc(columna) + 1
                        
                    # Agregar el hallazgo al identificador correspondiente en el diccionario
                    if (f'{identificador}({str(codigo).split(".", 1)[0]})') in hallazgos:
                        hallazgos[f'{identificador}({str(codigo).split(".", 1)[0]})'] += f', Hallazgo {numero_hallazgo}'
                    else:
                        hallazgos[f'{identificador}({str(codigo).split(".", 1)[0]})'] = f'Hallazgo {numero_hallazgo}'
            
        if not((f'{identificador}({str(codigo).split(".", 1)[0]})') in hallazgos):
            hallazgos[f'{identificador}({str(codigo).split(".", 1)[0]})'] = 'Sin hallazgos'

    errores_df = pd.DataFrame(list(errores.items()), columns=['ESTACION', 'ERRORES'])
    hallazgos_df = pd.DataFrame(list(hallazgos.items()), columns=['ESTACION', 'HALLAZGOS'])
    estado_df = pd.DataFrame(list(estado.items()), columns=['ESTACION', 'ESTADO'])
    codigo_df = pd.DataFrame(list(codigo_localizacion.items()), columns=['ESTACION', 'CODIGO LOCALIZACION'])

    uno_df = pd.DataFrame()
    uno_df = estado_df.merge(
        codigo_df,
        how='outer',
        on='ESTACION'
    )

    dos_df = pd.DataFrame()
    dos_df = uno_df.merge(
        errores_df,
        how='outer',
        on='ESTACION'
    )

    df_final = pd.DataFrame()
    df_final = dos_df.merge(
        hallazgos_df,
        how='outer',
        on='ESTACION'
    )

    df_final['ESTACION'] = df_final['ESTACION'].str.split('(').str[0]

    if nada is True:
        for row_index, row in df_final.iterrows():
            if row['ERRORES'] == 'Sin errores' and row['HALLAZGOS'] == 'Sin hallazgos':
                df_final = df_final.drop(index=row_index)
        st.dataframe(df_final, hide_index=True, use_container_width=True)
    elif df_final.empty:
        st.info('No se encontro ningún error o hallazgo')
    else:
        df_final = df_final.sort_values('ESTACION')
        st.dataframe(df_final, hide_index=True, use_container_width=True)
    
    if len(df_final) != 0:
        ruta = 'Verificador_Epocas_NSCL.xlsx'
        reporte(df_final, ruta)
    else:
        st.success('No se encontraron ni errores ni hallazgos')

# Funcion que se encarga de verificar los documentos asociados a cada estacion
def verificar_tres(df_tres):
    st.dataframe(df_tres, hide_index=True, use_container_width=True)

    st.divider()

    errores = {}
    estado = {}

    st.subheader('Codigos de error', help='El desplegable "Errores" contiene todas las comparaciones que evaluará el aplicativo en el reporte')
    with st.expander('Errores'):
        st.subheader('Errores:')

        st.info('1. La estación tiene una cantidad incorrecta de "Historicos de la estación", debería contar con dos "Historicos de la estación".')
        st.info('2. La estación tiene una cantidad incorrecta de "Pruebas del dataless", debería contar con solo una "Pruebas del dataless".')
        st.info('3. La estación tiene una cantidad incorrecta de "Archivo de respuesta", debería contar con solo un "Archivo de respuesta".')
        st.info('4. La estacion no cuenta con Informe de calidad de la señal.')
    
    # Selección de las estaciones a verificar
    st.subheader('Estaciones por verificar', help='Aquí se genera una vista previa del reporte con las incoherencias y errores de cada estación, el usuario puede elegir las estaciones que quiere verificar, en caso de no elegir ninguna el aplicativo verificara todas las estaciones disponibles.')
    def eleccion(datos):
        opcion = st.multiselect(
            'Seleccione las estaciones a verificar:',
            options=datos['ID ESTACION'].unique()
        ,key="mult_select_est")

        if len(opcion) == 0:
            respuesta = datos
        else:
            respuesta = datos[datos['ID ESTACION'].isin(opcion)]
        
        return respuesta
    respuesta = eleccion(df_tres)

    st.subheader('Agencia', help='Si no se selecciona ninguna agencia, el aplicativo tomará las estaciones que tengan como agencia "SERVICIO GEOLOGICO COLOMBIANO"')
    def agencia(respuesta):
        opciones = st.multiselect(
            'Seleccione las agencias a verificar:',
            options=respuesta['AGENCIA'].unique(),
            default=['SERVICIO GEOLOGICO COLOMBIANO']
        ,key="mult_select_agen")

        if len(opciones) == 0:
            respuesta = respuesta
        else:
            respuesta = respuesta[respuesta['AGENCIA'].isin(opciones)]
        
        return respuesta
    respuesta = agencia(respuesta)

    datos = respuesta.groupby(['ID ESTACION', 'TIPO ARCHIVO']).size().unstack(fill_value=0)

    for index, row in datos.iterrows():
        identificador = index

        if row['Histórico de la estación'] != 2:
            if identificador in errores:
                errores[identificador] += ', Error 1'
            else:
                errores[identificador] = 'Error 1'
        
        if row['Pruebas del Dataless'] != 1:
            if identificador in errores:
                errores[identificador] += ', Error 2'
            else:
                errores[identificador] = 'Error 2'
        
        if row['Archivo de respuesta'] != 1:
            if identificador in errores:
                errores[identificador] += ', Error 3'
            else:
                errores[identificador] = 'Error 3'
        
        if row['Informe de calidad de la señal de la estación'] < 1:
            if identificador in errores:
                errores[identificador] += ', Error 4'
            else:
                errores[identificador] = 'Error 4'
                
    for index_, row_ in respuesta.iterrows():
        identificador_ = row_['ID ESTACION']
        estado_estacion = row_['ESTADO']

        if identificador_ in estado:
            pass
        else:
            estado[identificador_] = estado_estacion
    
    errores_df = pd.DataFrame(list(errores.items()), columns=['ESTACION', 'ERRORES'])
    estado_df = pd.DataFrame(list(estado.items()), columns=['ESTACION','ESTADO'])

    df_final = pd.DataFrame()
    df_final = estado_df.merge(
        errores_df,
        how='outer',
        on='ESTACION'
    )

    if len(df_final) != 0:
        st.dataframe(df_final, hide_index=True, use_container_width=True)
        ruta = 'Verificador_Documentos.xlsx'
        reporte(errores_df, ruta)
    else:
        st.success('No se encontraron ni errores ni hallazgos')