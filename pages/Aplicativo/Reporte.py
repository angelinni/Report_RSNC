import streamlit as st
import openpyxl
import datetime
import os.path
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.drawing.image import Image
from io import BytesIO

# Generacion y descarga del reporte
def reporte(resultado, ruta):
    # Guarda el reporte en una ruta especifica
    resultado.to_excel(os.path.dirname(os.path.abspath(__file__))+'/Reportes/' + ruta, index=False, startrow=9)
    libro = openpyxl.load_workbook(os.path.dirname(os.path.abspath(__file__))+'/Reportes/' + ruta)
    hoja = libro.active

    imagen = Image('Imagenes/Plantilla.png')
    imagen.height = 140
    imagen.width = 300
    hoja.add_image(imagen, 'A1')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    fecha = datetime.datetime.now()
    hoja['A8'] = 'Fecha de creción:'
    hoja['B8'] = fecha
    hoja.cell(row=8, column=1).border = thin_border

    # Cambia el color del fondo de los titulos
    for celda in hoja[10]:
        celda.fill = PatternFill(start_color='8CA448', end_color='8CA448', fill_type='solid')
                
    # Ajusta el anchos de las columnas para los datos del reporte
    for columna in hoja.columns:
        max_length = 0
        column = columna[0].column_letter
        for celda in columna:
            try:
                if len(str(celda.value)) > max_length:
                    max_length = len(celda.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        hoja.column_dimensions[column].width = adjusted_width
    libro.save(os.path.dirname(os.path.abspath(__file__))+'/Reportes/' + ruta)
                
    # Permite descargar el reporte con el boton "Descargar Reporte"
    with open(os.path.dirname(os.path.abspath(__file__))+'/Reportes/' + ruta, 'rb') as file:
        contenido = file.read()
    st.download_button(label='Descargar Reporte', data=BytesIO(contenido), file_name=ruta)