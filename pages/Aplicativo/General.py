import streamlit as st
import os.path
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
#from Reporte import reporte
import openpyxl
import datetime
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.drawing.image import Image
from io import BytesIO




titulo = [
    'Estado de las estaciones',
    'Red de monitoreo de las estaciones',
    'Sistema de energía de las estaciones'
]
subtitulo = [
    'Estado',
    'Red de monitoreo',
    'Sistema de energía'
]
ayuda = [
    'Si no se selecciona ningún estado, el aplicativo no generará el gráfico correspondiente a este filtro',
    'Si no se selecciona ninguna red de monitoreo, el aplicativo no generará el gráfico correspondiente a este filtro',
    'Si no se selecciona ningún sistema de energía, el aplicativo no generará el gráfico correspondiente a este filtro'
]
columna = [
    'ESTADO',
    'RED MONITOREO',
    'SISTEMA DE ENERGIA'
]
seleccion = [
    'Seleccione los estados de las estaciones:',
    'Selecciones las redes de monitoreo de las estaciones:',
    'Selecciones las redes de monitoreo de las estaciones:'
]

def filtros(datos, codigo):
    st.subheader(titulo[codigo], help=ayuda[codigo])

    modificado = True
    datos_filtrados = datos.dropna(subset=[columna[codigo]])

    opciones = st.multiselect(
        seleccion[codigo],
        options = datos_filtrados[columna[codigo]].unique()
    )

    if len(opciones) == 0:
        respuesta = datos
        modificado = False
    elif len(opciones) == len(datos_filtrados[columna[codigo]].unique()):
        respuesta = datos_filtrados
    else:
        respuesta = datos_filtrados[datos_filtrados[columna[codigo]].isin(opciones)]

    return respuesta, modificado

def grafico(resultado, modificado, codigo):
    if modificado is True:
        st.subheader(subtitulo[codigo])

        datos_filtrados = resultado.groupby(columna[codigo]).size().reset_index(name='AGRUPADO')
        grafico = px.bar(
            datos_filtrados,
            title=False,
            x=columna[codigo],
            y='AGRUPADO',
            color=columna[codigo],
            text_auto=True,
            hover_data = {
                'AGRUPADO': False,
                'NUMERO DE ESTACIONES': datos_filtrados['AGRUPADO']
            }
        )

        grafico.update_yaxes(title='NUMERO DE ESTACIONES')
        grafico.update_xaxes(title=columna[codigo])
        grafico.update_traces(textposition='outside')

        st.plotly_chart(grafico, use_container_width=True)

### Generacion y descarga del reporte
def reporte(resultado, ruta):
    # Guarda el reporte en una ruta especifica
    resultado.to_excel(os.path.dirname(os.path.abspath(__file__))+'/Reportes/' + ruta, index=False, startrow=9)
    libro = openpyxl.load_workbook(os.path.dirname(os.path.abspath(__file__))+'/Reportes/' + ruta)
    hoja = libro.active

    imagen = Image(os.path.dirname(os.path.abspath(__file__))+'/Imagenes/Plantilla.png')
    imagen.height = 140
    imagen.width = 300
    hoja.add_image(imagen, 'A1')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    fecha = datetime.datetime.now()
    hoja['A8'] = 'Fecha de creción:'
    hoja['B8'] = fecha
    hoja.cell(row=8, column=1).border = thin_border

    # Cambia el color del fondo de los titulos
    for celda in hoja[10]:
        celda.fill = PatternFill(start_color='8CA448', end_color='8CA448', fill_type='solid')
                
    # Ajusta el anchos de las columnas para los datos del reporte
    for columna in hoja.columns:
        max_length = 0
        column = columna[0].column_letter
        for celda in columna:
            try:
                if len(str(celda.value)) > max_length:
                    max_length = len(celda.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        hoja.column_dimensions[column].width = adjusted_width
    libro.save(os.path.dirname(os.path.abspath(__file__))+'/Reportes/' + ruta)
                
    # Permite descargar el reporte con el boton "Descargar Reporte"
    with open(os.path.dirname(os.path.abspath(__file__))+'/Reportes/' + ruta, 'rb') as file:
        contenido = file.read()
    st.download_button(label='Descargar Reporte', data=BytesIO(contenido), file_name=ruta)


class Principal:

    def __init__(self, df):
        self.df = df
    
    # Vista de los datos cargados
    def visualizar(self):
        st.dataframe(self.df, hide_index=True, use_container_width=True)
    
    # Reporte general con filtros
    def seccion_uno(self):
        st.subheader('Reporte general con filtros', help='El aplicativo generará gráficos y un reporte teniendo en cuenta los filtros que se seleccionen. En caso de no seleccionar ningún filtro, el aplicativo no generará ningún gráfico')
        datos = self.df

        with st.expander('Filtros'):
            # Filtro y grafico por el estado de las estaciones
            codigo_uno = 0
            resultado_uno, modificado_uno = filtros(datos, codigo_uno)

            # Filtro y grafico por la red de monitoreo de las estaciones
            codigo_dos = 1
            resultado_dos, modificado_dos = filtros(resultado_uno, codigo_dos)

            # Filtro y grafico por el sistema de energia de las estaciones
            codigo_tres = 2
            resultado_tres, modificado_tres = filtros(resultado_dos, codigo_tres)
            
            col1, col2, col3, col4, col5 = st.columns(5)

            with col3:
                # Boton "Aplicar Filtros"
                boton = st.button('Aplicar Filtros', use_container_width=True)
        
        if boton:
            grafico(resultado_uno, modificado_uno, codigo_uno)
            grafico(resultado_dos, modificado_dos, codigo_dos)
            grafico(resultado_tres, modificado_tres, codigo_tres)

            st.dataframe(resultado_tres, hide_index=True, use_container_width=True)
            ruta = 'Reporte_General_Filtrado.xlsx'
            reporte(resultado_tres, ruta)
    
    # Estaciones instaladas por año
    def seccion_dos(self):
        st.subheader('Estaciones instaladas por año', help='El aplicativo generará el gráfico y un reporte teniendo en cuenta los filtros que se seleccionen. En caso de no seleccionar ningún filtro, el aplicativo generará un gráfico y un reporte con todos los datos dispoibles')
        datos = self.df

        def rango_fechas(datos):
            st.subheader('Rango de años', help='Permite modificar el rango de años a graficar. En caso de no realizar cambios en este rango, el aplicativo utilizará de forma predeterminada un rango que incluya todas las estaciones disponibles')

            datos_filtrados = datos.dropna(subset=['FECHA INSTALACION'])
            
            datos_filtrados = datos_filtrados.replace({"None":None})
            
            datos_filtrados['FECHA INSTALACION'] = pd.to_datetime(datos_filtrados['FECHA INSTALACION'])
            datos_filtrados['AÑO INSTALACION'] = datos_filtrados['FECHA INSTALACION'].dt.year
            datos_filtrados['FECHA INSTALACION'] = pd.to_datetime(datos_filtrados['FECHA INSTALACION']).dt.date
            
            

            año_inicio = datos_filtrados.sort_values(by=['AÑO INSTALACION'])
            año_fin = datos_filtrados.sort_values(by=['AÑO INSTALACION'])

            col1, col2 = st.columns(2)

            with col1:
                opcion_uno = st.selectbox(
                    'Seleccione el año inicial del rango:',
                    año_inicio['AÑO INSTALACION'].unique(),
                    index=0,
                    key=1
                )

            with col2:
                opcion_dos = st.selectbox(
                    'Seleccione el año final del rango:',
                    año_fin['AÑO INSTALACION'].unique(),
                    index=0,
                    key=2
                )

            respuesta = datos_filtrados[((datos_filtrados['AÑO INSTALACION'] >= opcion_uno) & (datos_filtrados['AÑO INSTALACION'] <= opcion_dos))]
            
            return respuesta
        resultado = rango_fechas(datos)
        
        def agencia(resultado):
            st.subheader('Agencia', help='Permite filtrar por la agencia de las estaciones. En caso de no realizar cambios en este filtro, el aplicativo utilizará de forma predeterminada todas las agencias de las estaciones disponibles')

            datos_filtrados = resultado.dropna(subset=['AGENCIA'])
            
            opciones = st.multiselect(
                'Seleccione las agencias de las estaciones:',
                options = datos_filtrados['AGENCIA'].unique(),
                key='uno',
                default='SERVICIO GEOLOGICO COLOMBIANO'
            )

            if len(opciones) == 0:
                respuesta = resultado
            elif len(opciones) == len(datos_filtrados['AGENCIA'].unique()):
                respuesta = datos_filtrados
            else:
                respuesta = datos_filtrados[datos_filtrados['AGENCIA'].isin(opciones)]
            
            return respuesta
        resultado = agencia(resultado)

        def grafico_instaladas(resultado):
            datos_filtrados = resultado.groupby('AÑO INSTALACION').size().reset_index(name='AÑO INSTALACION AGRUPADO')

            grafico = px.bar(
                datos_filtrados,
                title = False,
                x='AÑO INSTALACION',
                y='AÑO INSTALACION AGRUPADO',
                text_auto=True,
                hover_data={
                    'AÑO INSTALACION AGRUPADO': False,
                    'NUMERO DE ESTACIONES': datos_filtrados['AÑO INSTALACION AGRUPADO']
                }
            )

            grafico.update_yaxes(title='NUMERO DE ESTACIONES')
            grafico.update_xaxes(tickmode='linear')
            grafico.update_traces(textposition='outside')

            st.plotly_chart(grafico, use_container_width=True)
        grafico_instaladas(resultado)

        st.dataframe(resultado, hide_index=True, use_container_width=True)
        ruta = 'Reporte_Estaciones_Instaladas.xlsx'
        reporte(resultado, ruta)
    
    # Estaciones retiradas por año
    def seccion_tres(self):
        st.subheader('Estaciones retiradas por año', help='El aplicativo generará el gráfico y un reporte teniendo en cuenta los filtros que se seleccionen. En caso de no seleccionar ningún filtro, el aplicativo generará un gráfico y un reporte con todos los datos dispoibles')
        datos = self.df

        def rango_fechas(datos):
            st.subheader('Rango de años', help='Permite modificar el rango de años a graficar. En caso de no realizar cambios en este rango, el aplicativo utilizará de forma predeterminada un rango que incluya todas las estaciones disponibles')

            datos_filtrados = datos.dropna(subset=['FECHA RETIRO'])
            datos_filtrados = datos_filtrados.replace({"None":None})
            datos_filtrados['FECHA RETIRO'] = pd.to_datetime(datos_filtrados['FECHA RETIRO'])
            datos_filtrados['AÑO RETIRO'] = datos_filtrados['FECHA RETIRO'].dt.year
            datos_filtrados['FECHA RETIRO'] = pd.to_datetime(datos_filtrados['FECHA RETIRO']).dt.date

            año_inicio = datos_filtrados.sort_values(by=['AÑO RETIRO'])
            año_fin = datos_filtrados.sort_values(by=['AÑO RETIRO'])

            col1, col2 = st.columns(2)

            with col1:
                opcion_uno = st.selectbox(
                    'Seleccione el año inicial del rango:',
                    año_inicio['AÑO RETIRO'].unique(),
                    index=0,
                    key=3
                )

            with col2:
                opcion_dos = st.selectbox(
                    'Seleccione el año final del rango:',
                    año_fin['AÑO RETIRO'].unique(),
                    index=0,
                    key=4
                )

            respuesta = datos_filtrados[((datos_filtrados['AÑO RETIRO'] >= opcion_uno) & (datos_filtrados['AÑO RETIRO'] <= opcion_dos))]
            return respuesta
        resultado = rango_fechas(datos)

        def agencia(resultado):
            st.subheader('Agencia', help='Permite filtrar por la agencia de las estaciones. En caso de no realizar cambios en este filtro, el aplicativo utilizará de forma predeterminada todas las agencias de las estaciones disponibles')

            datos_filtrados = resultado.dropna(subset=['AGENCIA'])
            opciones = st.multiselect(
                'Seleccione las agencias de las estaciones:',
                options = datos_filtrados['AGENCIA'].unique(),
                key='dos',
                default='SERVICIO GEOLOGICO COLOMBIANO'
            )

            if len(opciones) == 0:
                respuesta = resultado
            elif len(opciones) == len(datos_filtrados['AGENCIA'].unique()):
                respuesta = datos_filtrados
            else:
                respuesta = datos_filtrados[datos_filtrados['AGENCIA'].isin(opciones)]
            
            return respuesta
        resultado = agencia(resultado)

        def grafico_instaladas(resultado):
            datos_filtrados = resultado.groupby('AÑO RETIRO').size().reset_index(name='AÑO RETIRO AGRUPADO')

            grafico = px.bar(
                datos_filtrados,
                title = False,
                x='AÑO RETIRO',
                y='AÑO RETIRO AGRUPADO',
                text_auto=True,
                hover_data={
                    'AÑO RETIRO AGRUPADO': False,
                    'NUMERO DE ESTACIONES': datos_filtrados['AÑO RETIRO AGRUPADO']
                }
            )

            grafico.update_yaxes(title='NUMERO DE ESTACIONES')
            grafico.update_xaxes(tickmode='linear')
            grafico.update_traces(textposition='outside')

            st.plotly_chart(grafico, use_container_width=True)
        grafico_instaladas(resultado)

        st.dataframe(resultado, hide_index=True, use_container_width=True)
        ruta = 'Reporte_Estaciones_Retiradas.xlsx'
        reporte(resultado, ruta)
    
    # Acumulado de estaciones por año
    def seccion_cuatro(self):
        st.subheader('Acumulado de estaciones por año', help='Aquí se muestra el acumulado de las estaciones instaladas y retiradas por año, en caso de nececitar visualizar solo uno de las 2 opciones podemos deshabilitarlas dando clic en la leyenda de la derecha.')
        datos = self.df

        def acumulado(datos):
            datos_filtrados = datos

            datos_filtrados = datos_filtrados.replace({"None":None})
            datos_filtrados['FECHA INSTALACION'] = pd.to_datetime(datos_filtrados['FECHA INSTALACION'])
            datos_filtrados['FECHA RETIRO'] = pd.to_datetime(datos_filtrados['FECHA RETIRO'])

            # Creamos 2 columnas en las cuales guardamos solo el años tanto de instalacion como de retiro de cada estacion
            datos_filtrados['AÑO INSTALACION'] = datos_filtrados['FECHA INSTALACION'].dt.year
            datos_filtrados['AÑO RETIRO'] = datos_filtrados['FECHA RETIRO'].dt.year

            # En 2 nuevos dataframes agrupamos el año de cada estacion para obtener la cantidad de estaciones que hubo en dicho año
            df_instaladas = datos_filtrados.groupby('AÑO INSTALACION').size().reset_index(name='ESTACIONES INSTALADAS')
            df_retiradas = datos_filtrados.groupby('AÑO RETIRO').size().reset_index(name='ESTACIONES RETIRADAS')
            
            # Agregamos 2 columnas adicionales a los 2 dataframes que creamos donde guardaremos el acumulado de las estaciones año a año
            df_instaladas['ACUMULADO ESTACIONES INSTALADAS'] = df_instaladas['ESTACIONES INSTALADAS'].cumsum()
            df_retiradas['ACUMULADO ESTACIONES RETIRADAS'] = df_retiradas['ESTACIONES RETIRADAS'].cumsum()
            
            # Crea la grafica principal
            grafico = go.Figure()

            # Agregar la primera gráfica donde mostraremos las estaciones instaladas
            grafico.add_trace(go.Bar(
                x=df_instaladas['AÑO INSTALACION'],
                y=df_instaladas['ACUMULADO ESTACIONES INSTALADAS'],
                name='INSTALADAS',
                hovertemplate='AÑO: %{x}<br>ESTACIONES INSTALADAS: %{y}',
                text=df_instaladas['ACUMULADO ESTACIONES INSTALADAS'],
            ))

            # Agregar la segunda gráfica donde mostraremos las estaciones retiradas
            grafico.add_trace(go.Bar(
                x=df_retiradas['AÑO RETIRO'],
                y=df_retiradas['ACUMULADO ESTACIONES RETIRADAS'],
                name='RETIRADAS',
                hovertemplate='AÑO: %{x}<br>ESTACIONES RETIRADAS: %{y}',
                text=df_retiradas['ACUMULADO ESTACIONES RETIRADAS']
            ))

            # Personaliza los titulos de la grafica
            grafico.update_layout(
                xaxis_title='AÑO',
                yaxis_title='NUMERO DE ESTACIONES',
                barmode='group'
            )

            grafico.update_xaxes(tickmode='linear')
            grafico.update_traces(textposition='outside')

            st.plotly_chart(grafico, use_container_width=True)
        acumulado(datos)
