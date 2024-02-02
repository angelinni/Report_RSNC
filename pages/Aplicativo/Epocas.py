import streamlit as st
import os.path
import pandas as pd
import datetime
import plotly.express as px
#from Reporte import reporte
import openpyxl
import datetime
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.drawing.image import Image
from io import BytesIO

titulo = [
    'Tipo de almacenamiento de las estaciones',
    'Condición de instalación de las estaciones',
    'Tipo de transmisión de las estaciones',
    'Tipo de descarga de las estaciones',
    'Tipo de alcance de las estaciones',
    'Tipo de estación',
    'Tipo de adquisición de las estaciones'
]
subtitulo = [
    'Tipo de almacenamiento',
    'Condición de instalación',
    'Tipo de transmisión',
    'Tipo de descarga',
    'Tipo de alcance',
    'Tipo de estación',
    'Tipo de adquisición'
]
ayuda = [
    'Si no se selecciona ningún tipo de almacenamiento, el aplicativo no generará el gráfico correspondiente a este filtro',
    'Si no se selecciona ninguna condición de instalación, el aplicativo no generará el gráfico correspondiente a este filtro',
    'Si no se selecciona ninguna tipo de transmisión, el aplicativo no generará el gráfico correspondiente a este filtro',
    'Si no se selecciona ningún tipo de descarga, el aplicativo no generará el gráfico correspondiente a este filtro',
    'Si no se selecciona ningún tipo de alcance, el aplicativo no generará el gráfico correspondiente a este filtro',
    'Si no se selecciona ningún tipo de estación, el aplicativo no generará el gráfico correspondiente a este filtro',
    'Si no se selecciona ningún tipo de adquisición, el aplicativo no generará el gráfico correspondiente a este filtro'
]
columna = [
    'TIPO ALMACENAMIENTO',
    'CONDICIÓN DE INSTALACION',
    'TIPO DE TRANMISION',
    'TIPO DESCARGA',
    'TIPO DE ALCANCE',
    'TIPO DE ESTACION',
    'TIPO DE ADQUISICION'
]
seleccion = [
    'Selecciones los tipos de almacenamiento de las estaciones:',
    'Seleccione las condiciones de instalación de las estaciones:',
    'Seleccione los tipos de transmisión de las estaciones:',
    'Seleccione los tipos de descarga de las estaciones:',
    'Seleccione los tipos de alcance de las estaciones:',
    'Seleccione los tipos de estacion:',
    'Seleccione los tipos de adquisición de las estaciones:'
]

def filtros(datos, codigo):
    st.subheader(titulo[codigo], help=ayuda[codigo])

    modificado = True
    datos_filtrados = datos.dropna(subset=[columna[codigo]])

    opciones = st.multiselect(
        seleccion[codigo],
        options = datos_filtrados[columna[codigo]].unique()
    )

    if len(opciones) == 0:
        respuesta = datos
        modificado = False
    elif len(opciones) == len(datos_filtrados[columna[codigo]].unique()):
        respuesta = datos_filtrados
    else:
        respuesta = datos_filtrados[datos_filtrados[columna[codigo]].isin(opciones)]

    return respuesta, modificado

def grafico(resultado, modificado, codigo):
    if modificado is True:
        st.subheader(subtitulo[codigo])

        datos_filtrados = resultado.groupby(columna[codigo]).size().reset_index(name='AGRUPADO')
        grafico = px.bar(
            datos_filtrados,
            title=False,
            x=columna[codigo],
            y='AGRUPADO',
            color=columna[codigo],
            text_auto=True,
            hover_data = {
                'AGRUPADO': False,
                'NUMERO DE ESTACIONES': datos_filtrados['AGRUPADO']
            }
        )

        grafico.update_yaxes(title='NUMERO DE ESTACIONES')
        grafico.update_xaxes(title=columna[codigo])
        grafico.update_traces(textposition='outside')

        st.plotly_chart(grafico, use_container_width=True)

### Generacion y descarga del reporte
def reporte(resultado, ruta):
    # Guarda el reporte en una ruta especifica
    resultado.to_excel(os.path.dirname(os.path.abspath(__file__))+'/Reportes/' + ruta, index=False, startrow=9)
    libro = openpyxl.load_workbook(os.path.dirname(os.path.abspath(__file__))+'/Reportes/' + ruta)
    hoja = libro.active

    imagen = Image(os.path.dirname(os.path.abspath(__file__))+'/Imagenes/Plantilla.png')
    imagen.height = 140
    imagen.width = 300
    hoja.add_image(imagen, 'A1')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    fecha = datetime.datetime.now()
    hoja['A8'] = 'Fecha de creción:'
    hoja['B8'] = fecha
    hoja.cell(row=8, column=1).border = thin_border

    # Cambia el color del fondo de los titulos
    for celda in hoja[10]:
        celda.fill = PatternFill(start_color='8CA448', end_color='8CA448', fill_type='solid')
                
    # Ajusta el anchos de las columnas para los datos del reporte
    for columna in hoja.columns:
        max_length = 0
        column = columna[0].column_letter
        for celda in columna:
            try:
                if len(str(celda.value)) > max_length:
                    max_length = len(celda.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        hoja.column_dimensions[column].width = adjusted_width
    libro.save(os.path.dirname(os.path.abspath(__file__))+'/Reportes/' + ruta)
                
    # Permite descargar el reporte con el boton "Descargar Reporte"
    with open(os.path.dirname(os.path.abspath(__file__))+'/Reportes/' + ruta, 'rb') as file:
        contenido = file.read()
    st.download_button(label='Descargar Reporte', data=BytesIO(contenido), file_name=ruta)


class PrincipalEp:

    def __init__(self, df):
        self.df = df
    
    def visualizar(self):
        st.dataframe(self.df, hide_index=True, use_container_width=True)
    
    def seccion_uno(self):
        st.subheader('Reporte épocas con filtros', help='El aplicativo generará gráficos y un reporte teniendo en cuenta los filtros que se seleccionen. En caso de no seleccionar ningún filtro, el aplicativo no generará ningún gráfico.')
        datos = self.df
        
        # Desplegable que contiene todos los filtros de esta sección
        with st.expander('Filtros'):
            
            # Filtro por el tipo de almacenamiento de las estaciones
            codigo_uno = 0
            resultado, modificado_uno = filtros(datos, codigo_uno)

            # Filtro por la condición de instalación de las estaciones
            codigo_dos = 1
            resultado, modificado_dos = filtros(resultado, codigo_dos)

            # Filtro por el tipo de transmisión de las estaciones
            codigo_tres = 2
            resultado, modificado_tres = filtros(resultado, codigo_tres)

            # Filtro por el tipo de descarga de las estaciones
            codigo_cuatro = 3
            resultado, modificado_cuatro = filtros(resultado, codigo_cuatro)

            # Filtro por el tipo de alcance de las estaciones
            codigo_cinco = 4
            resultado, modificado_cinco = filtros(resultado, codigo_cinco)

            # Filtro por el tipo de estación
            codigo_seis = 5
            resultado, modificado_seis = filtros(resultado, codigo_seis)

            # Filtro por el tipo de adquisición de las estaciones
            codigo_siete = 6
            resultado, modificado_siete = filtros(resultado, codigo_siete)

            boton = st.button('Aplicar Filtros')
        
        # Gráficos y reporte a generar al pulsar el botón
        if boton:
            grafico(resultado, modificado_uno, codigo_uno)
            grafico(resultado, modificado_dos, codigo_dos)
            grafico(resultado, modificado_tres, codigo_tres)
            grafico(resultado, modificado_cuatro, codigo_cuatro)
            grafico(resultado, modificado_cinco, codigo_cinco)
            grafico(resultado, modificado_seis, codigo_seis)
            grafico(resultado, modificado_siete, codigo_siete)

            st.dataframe(resultado, hide_index=True, use_container_width=True)
            ruta = 'Reporte_Epocas_Filtrado.xlsx'
            reporte(resultado, ruta)
    
    def seccion_dos(self):
        st.subheader('Reporte tipo de adquisición', help='El aplicativo generará un gráfico teniendo en cuenta la estación seleccionada. En caso de no seleccionar ninguna estación, el aplicativo tomara por defecto la primera estación disponible.')
        datos = self.df

        actual = datetime.datetime.now()
        actual_dos = actual.strftime("%Y-%m-%d")
        
        datos['FECHA INICIO'] = datos['FECHA INICIO'].str.split(' ').str[0]
        datos['FECHA FIN'] = datos['FECHA FIN'].str.split(' ').str[0]
        datos.loc[datos['FECHA FIN'] > actual_dos, 'FECHA FIN'] = actual_dos

        # Filtro por el tipo de adquisición de las estaciones
        def adquisicion(datos):
            datos_filtrados = datos.dropna(subset=['TIPO DE ADQUISICION'])

            opciones = st.selectbox(
                'Seleccione la estacion a gráficar:',
                options = datos_filtrados['ID ESTACION'].unique()
            )

            respuesta = datos_filtrados[datos_filtrados['ID ESTACION'] == opciones]
            return respuesta
        resultado = adquisicion(datos)

        st.dataframe(resultado, hide_index=True, use_container_width=True)

        def grafico_adquisicion(resultado):
            grafico = px.timeline(
                resultado,
                title=False,
                x_start='FECHA INICIO',
                x_end='FECHA FIN',
                y='CODIGO LOCALIZACION',
                pattern_shape='TIPO DE ADQUISICION'
            )

            grafico.update_yaxes(
                title='CODIGO DE LOCALIZACION',
                type='category',
            )
            
            grafico.update_xaxes(
                title='AÑO DE INSTALACION',
                tickmode='array',
                tickvals = resultado['FECHA INICIO'],
                showgrid=True,
                gridcolor='gray',
            )

            grafico.update_traces(textposition='outside')

            st.plotly_chart(grafico, use_container_width=True)
        grafico_adquisicion(resultado)

    def seccion_tres(self):
        st.subheader('Reporte cambio tipo de adquisición', help='El aplicativo generará un gráfico y un reporte teniendo en cuenta los filtros que se seleccionen. En caso de no seleccionar ningún filtro, el aplicativo graficará todas las opciones disponibles.')

        with st.expander('Filtros'):
            datos = self.df

            def seleccion_estaciones(datos):
                st.subheader('Estaciones por graficar', help='Aquí puedes modificar el gráfico y el reporte que aparecen al final, seleccionando una o más estaciones, en caso de no modificar el filtro el aplicativo tomara todas las estaciones disponibles.')

                datos_filtrados = datos

                opciones = st.multiselect(
                    'Seleccione las estaciones a graficar:',
                    datos_filtrados['ID ESTACION'].unique(),
                )

                if len(opciones) == 0:
                    respuesta = datos
                elif len(opciones) == len(datos_filtrados['ID ESTACION'].unique()):
                    respuesta = datos_filtrados
                else:
                    respuesta = datos_filtrados[datos_filtrados['ID ESTACION'].isin(opciones)]
                
                return respuesta
            resultado = seleccion_estaciones(datos)

            def seleccion_adquisicion(resultado):
                st.subheader('Tipo de adquisición a graficar', help='Aquí puedes modificar el gráfico y el reporte que aparecen al final, seleccionando el tipo de adquisición, en caso de no modificar el filtro el aplicativo tomara todos los tipos de adquisición disponibles.')

                datos_filtrados = resultado.dropna(subset=['TIPO DE ADQUISICION'])
                datos_filtrados = datos_filtrados.sort_values(by=['ID ESTACION', 'CODIGO LOCALIZACION', 'FECHA INICIO'])

                datos_filtrados['CAMBIO ADQUISICION'] = (
                    (datos_filtrados['TIPO DE ADQUISICION'] != datos_filtrados['TIPO DE ADQUISICION'].shift())
                    &
                    (datos_filtrados['ID ESTACION'] == datos_filtrados['ID ESTACION'].shift())
                    &
                    (datos_filtrados['CODIGO LOCALIZACION'] == datos_filtrados['CODIGO LOCALIZACION'].shift())
                )

                df_final = datos_filtrados[(datos_filtrados['CAMBIO ADQUISICION'])]

                opciones = st.multiselect(
                    'Seleccione los tipos de adquisición a graficar:',
                    df_final['TIPO DE ADQUISICION'].unique()
                )

                if len(opciones) == 0:
                    respuesta = df_final
                elif len(opciones) == len(df_final['TIPO DE ADQUISICION'].unique()):
                    respuesta = df_final
                else:
                    respuesta = df_final[df_final['TIPO DE ADQUISICION'].isin(opciones)]

                return respuesta
            resultado = seleccion_adquisicion(resultado)

            def reporte_cambio_adquisicion(resultado):
                st.subheader('Rango de fechas a graficar', help='Aquí puedes modificar el gráfico y el reporte que aparecen al final, seleccionando el rango de fechas que se quiere graficar, en caso de no modificar el filtro el aplicativo tomara un rango de fechas que incluya todas las estaciones.')

                df_final = resultado

                fecha_minima = pd.to_datetime(df_final['FECHA INICIO']).min()
                fecha_maxima = pd.to_datetime(df_final['FECHA FIN']).max()
                print("### Aqui", type(fecha_maxima))

                col1, col2 = st.columns(2)

                # Le pedimos al usuario seleccionar la fecha de inicio del rango que usaremos
                with col1:
                    fecha_inicio = st.date_input("Ingrese la fecha de inicio", datetime.date(1993, 6, 1), 
                                                min_value=datetime.date(1993, 1, 1),
                                                max_value=datetime.date(2035, 1, 1))
                    fecha_inicio = fecha_inicio.strftime("%Y-%m-%d")

                # Le pedimos al usuario seleccionar la fecha de fin del rango que usaremos
                with col2:
                    fecha_fin = st.date_input("Ingrese la fecha de fin", datetime.date(1993, 6, 1), 
                                            min_value=datetime.date(1993, 1, 1),
                                            max_value=datetime.date(2035, 1, 1))
                    fecha_fin = fecha_fin.strftime("%Y-%m-%d")

                df_final = df_final[(df_final['FECHA INICIO'] >= fecha_inicio) & (df_final['FECHA INICIO'] <= fecha_fin)]

                return df_final
            resultado = reporte_cambio_adquisicion(resultado)

            boton = st.button('Aplicar Filtros', key=2)

        if boton:
            def grafico(resultado):
                grafico = px.timeline(
                    resultado,
                    title=False,
                    x_start='FECHA INICIO',
                    x_end='FECHA FIN',
                    y='ID ESTACION',
                    pattern_shape='TIPO DE ADQUISICION',
                )

                grafico.update_yaxes(
                    title='NOMBRE DE LA ESTACION',
                    type="category",
                    autorange='reversed',
                    categoryarray=resultado['ID ESTACION']
                )

                grafico.update_xaxes(
                    title='AÑO DE INSTALACION',
                    tickmode='array',
                    tickvals = resultado['FECHA INICIO'],
                    showgrid=False,
                    gridcolor='gray',
                )

                st.subheader('Gráfico y reporte')

                grafico.update_traces(textposition='outside')

                st.plotly_chart(grafico, use_container_width=True)
            grafico(resultado)

            st.dataframe(resultado, hide_index=True, use_container_width=True)
            ruta = 'Reporte_Epocas_Cambio_Adquisicion.xlsx'
            reporte(resultado, ruta)